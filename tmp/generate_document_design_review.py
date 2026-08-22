import os
import sys
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.dev')

import django

django.setup()

from django.db import transaction
from django.test import RequestFactory
from django.test.runner import DiscoverRunner
from pypdf import PdfReader
from openpyxl import load_workbook

from controles.models import (
    EquipamentoLocadoCatalogo,
    FaturamentoDireto,
    LocacaoEquipamento,
    LocadoraEquipamento,
    OrcamentoRadarObra,
)
from controles.views import _radar_obras_pdf, relatorio_locacoes_equipamentos_pdf
from documentos.excel import ExcelReportBuilder
from documentos.formatting import format_money_br
from empresas.models import Empresa
from financeiro.views import (
    FinanceiroFiltroForm,
    _financial_report_pdf,
    _xlsx_relatorio_financeiro,
)
from medicoes.forms import RelatorioMedicoesForm
from medicoes.models import (
    Empreiteiro,
    FaturamentoDiretoMedicao,
    ItemMedicaoConstrutora,
    ItemMedicaoEmpreiteiro,
    ItemOrcamentoMedicao,
    MedicaoConstrutora,
    MedicaoEmpreiteiro,
    OrcamentoMedicao,
)
from medicoes.views import (
    _itens_medicao_construtora_com_grupos,
    _pdf_medicao_construtora,
    _pdf_medicao_empreiteiro,
    _pdf_relatorio_medicoes,
    _pdf_saldo_contratual,
    _totais_relatorio_medicoes,
    _saldo_contratual_construtora,
    _xlsx_medicao,
    _xlsx_relatorio_medicoes,
    _xlsx_saldo_contratual,
)
from obras.models import Obra


OUT = Path('tmp/document_design_review')
OUT.mkdir(parents=True, exist_ok=True)


def write(path, content):
    path.write_bytes(content)
    return path


def pdf_pages(path):
    return len(PdfReader(str(path)).pages)


def excel_flags(path):
    wb = load_workbook(path)
    ws = wb.active
    money = date_cell = percent = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            fmt = str(cell.number_format or '')
            if 'R$' in fmt:
                money = True
            if 'd' in fmt.lower() and 'y' in fmt.lower():
                date_cell = True
            if '%' in fmt:
                percent = True
    return {
        'freeze': bool(ws.freeze_panes),
        'filter': bool(ws.auto_filter.ref),
        'money': money,
        'date': date_cell,
        'percent': percent,
        'a4': str(ws.page_setup.paperSize) == str(ws.PAPERSIZE_A4),
        'fit_width': ws.page_setup.fitToWidth == 1,
        'orientation': ws.page_setup.orientation,
        'footer': bool(ws.oddFooter.center.text or ws.oddFooter.right.text),
        'merged': len(ws.merged_cells.ranges),
    }


def make_request(empresa):
    request = RequestFactory().get('/')
    request.empresa = empresa
    return request


def create_empresa(nome, slug, prim, sec):
    return Empresa.objects.create(
        nome=nome,
        razao_social=nome,
        nome_fantasia=nome,
        cnpj='00.000.000/0001-00',
        endereco='Rua Demonstracao, 100',
        cidade='Porto Alegre',
        estado='RS',
        cep='90000-000',
        telefone='(51) 3000-0000',
        email='teste@example.com',
        slug=slug,
        cor_primaria=prim,
        cor_secundaria=sec,
        texto_rodape=f'{nome} - documento sintetico para QA visual',
    )


def create_measurement_data(empresa):
    obra = Obra.objects.create(
        empresa=empresa,
        nome_obra='Obra Demonstracao Sao Jose',
        cliente='Cliente Joao Demonstracao',
        responsavel='Engenharia QA',
        valor_contrato=Decimal('1234567.89'),
        projecao_despesa=Decimal('456789.12'),
    )
    orcamento = OrcamentoMedicao.objects.create(
        obra=obra,
        nome='Planilha demonstracao - Medicao da Construcao',
        tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
    )
    itens = []
    ordem = 1
    for grupo_idx, grupo in enumerate(['SERVICOS PRELIMINARES', 'TERRAPLENAGEM', 'DRENAGEM PLUVIAL', 'PAVIMENTACAO']):
        ItemOrcamentoMedicao.objects.create(
            orcamento=orcamento,
            tipo=ItemOrcamentoMedicao.TIPO_GRUPO,
            ordem=ordem,
            item=str(grupo_idx + 1),
            descricao=grupo,
        )
        ordem += 1
        for i in range(1, 11):
            itens.append(
                ItemOrcamentoMedicao.objects.create(
                    orcamento=orcamento,
                    tipo=ItemOrcamentoMedicao.TIPO_ITEM,
                    ordem=ordem,
                    item=f'{grupo_idx + 1}.{i}',
                    descricao=(
                        'Execucao de servico de demonstracao com descricao propositalmente extensa '
                        'para validar quebra automatica de texto, altura da linha e comportamento '
                        'da tabela durante a paginacao com acentuacao: Medicao, Construção, Serviço, Mão de obra.'
                    )
                    if i in {3, 7}
                    else f'Servico demonstracao {grupo_idx + 1}.{i}',
                    unidade='m2' if i % 2 else 'm3',
                    quantidade=Decimal('100.0000') + Decimal(i * 3),
                    preco_unitario_material=Decimal('12.3500') if i == 1 else Decimal(i * 7),
                    preco_unitario_mao_obra=Decimal('25.2500') if i % 3 else Decimal('1234.5600'),
                    preco_unitario_equipamentos=Decimal('0.0000') if i % 4 else Decimal('45.6700'),
                )
            )
            ordem += 1

    medicao_1 = MedicaoConstrutora.objects.create(
        orcamento=orcamento,
        numero=1,
        periodo_inicio=date(2026, 6, 1),
        periodo_fim=date(2026, 6, 30),
        data_medicao=date(2026, 6, 30),
    )
    for idx, item in enumerate(itens[:20], start=1):
        ItemMedicaoConstrutora.objects.create(
            medicao=medicao_1,
            item_orcamento=item,
            quantidade_periodo=Decimal(idx % 5 + 1),
        )

    medicao_2 = MedicaoConstrutora.objects.create(
        orcamento=orcamento,
        numero=2,
        periodo_inicio=date(2026, 7, 1),
        periodo_fim=date(2026, 7, 31),
        data_medicao=date(2026, 7, 31),
        retencao_tecnica_percentual=Decimal('5.2500'),
        issqn_percentual=Decimal('3.0000'),
        inss_percentual=Decimal('11.0000'),
        desconto_adicional=Decimal('1234.56'),
        desconto_adicional_reduz_base_nf=True,
    )
    for idx, item in enumerate(itens, start=1):
        ItemMedicaoConstrutora.objects.create(
            medicao=medicao_2,
            item_orcamento=item,
            quantidade_periodo=Decimal((idx % 6) + 1),
        )
    for i in range(1, 7):
        fd = FaturamentoDireto.objects.create(
            obra=obra,
            data_lancamento=date(2026, 7, i),
            numero_nf=f'FD-{i:03d}',
            numero_ordem_compra=f'OC-{i:03d}/2026',
            empresa_comprou='Empresa compradora de demonstracao',
            valor_nota=Decimal('1234.56') * i,
            descricao='Compra direta de material para abatimento em medicao',
            vencimento_boleto='30/07/2026',
            medicao_desconto='Medicao 02',
        )
        FaturamentoDiretoMedicao.objects.create(
            medicao=medicao_2,
            faturamento_direto=fd,
            percentual_descontado=Decimal('50.0000') if i % 2 else Decimal('100.0000'),
        )

    empreiteiro = Empreiteiro.objects.create(
        empresa=empresa,
        nome='Empreiteiro Joao Servicos',
        cpf_cnpj='000.000.000-00',
        pix='pix-demonstracao@example.com',
    )
    med_emp = MedicaoEmpreiteiro.objects.create(
        tipo=MedicaoEmpreiteiro.TIPO_CUMULATIVA,
        empresa=empresa,
        obra=obra,
        orcamento=orcamento,
        empreiteiro_cadastro=empreiteiro,
        numero=3,
        periodo_inicio=date(2026, 7, 1),
        periodo_fim=date(2026, 7, 31),
        data_medicao=date(2026, 7, 31),
        retencao_tecnica=Decimal('250.00'),
        desconto_adicional=Decimal('12.35'),
    )
    med_emp.save()
    for idx, item in enumerate(itens[:18], start=1):
        ItemMedicaoEmpreiteiro.objects.create(
            medicao=med_emp,
            item_orcamento=item,
            quantidade_periodo=Decimal((idx % 4) + 1),
        )
    return obra, orcamento, medicao_2, med_emp


def create_locacoes(empresa, obra):
    locadora = LocadoraEquipamento.objects.create(empresa=empresa, nome='Locadora Demonstracao')
    equipamentos = [
        EquipamentoLocadoCatalogo.objects.create(nome=f'Equipamento demonstracao {i}', categoria='Operacional')
        for i in range(1, 8)
    ]
    for i in range(60):
        LocacaoEquipamento.objects.create(
            equipamento=equipamentos[i % len(equipamentos)],
            locadora=locadora,
            obra=obra,
            data_locacao=date(2026, 1, 1) + timedelta(days=i),
            status=['locado', 'aguardando_entrega', 'retirada_solicitada', 'retirado'][i % 4],
            quantidade=(i % 5) + 1,
            data_retirada=date(2026, 3, 1) + timedelta(days=i) if i % 4 == 3 else None,
            observacoes='Observacao longa com acentuacao: execução, medição, construção e São José.',
        )


def create_radar(empresa):
    orcamentos = []
    for i in range(55):
        orcamentos.append(
            OrcamentoRadarObra.objects.create(
                empresa=empresa,
                numero=f'{i + 1:03d}/2026',
                cliente=f'Cliente Demonstracao {i + 1}',
                descricao='Proposta com descricao longa para validar wrap e relatório gerencial em A4.',
                data_orcamento=date(2026, 1, 1) + timedelta(days=i),
                situacao=['aguardando_resposta', 'em_revisao', 'fechada', 'nao_foi_para_frente'][i % 4],
                temperatura=(i % 5) + 1,
                valor_estimado=[Decimal('0'), Decimal('12.35'), Decimal('1234.56'), Decimal('123456.78'), Decimal('1234567.89')][i % 5],
                responsavel='Responsavel QA',
            )
        )
    return orcamentos


def financial_events():
    valores = [Decimal('0'), Decimal('12.35'), Decimal('1234.56'), Decimal('123456.78'), Decimal('1234567.89')]
    eventos = []
    for i in range(70):
        tipo = 'Receber' if i % 2 == 0 else 'Pagar'
        status = 'Recebido' if i % 4 == 0 else ('Pago' if i % 4 == 1 else 'Em aberto')
        eventos.append(
            {
                'tipo': tipo,
                'data': date(2026, 2, 1) + timedelta(days=i),
                'descricao': 'Lancamento financeiro sintetico com descricao longa e acentuacao: medição, execução, João e São José.',
                'pessoa': 'Fornecedor Teste' if tipo == 'Pagar' else 'Cliente Teste',
                'obra': 'Obra Demonstracao Sao Jose',
                'centro_custo': 'Centro demonstracao',
                'status': status,
                'valor': valores[i % len(valores)],
            }
        )
    resumo = {
        'total_receber_aberto': Decimal('1234567.89'),
        'total_pagar_aberto': Decimal('123456.78'),
        'saldo_com_previsoes': Decimal('1111111.11'),
        'saldo_realizado': Decimal('987654.32'),
    }
    return eventos, resumo


def medicao_report_lines(med_construtora, med_empreiteiro):
    linhas = []
    for i in range(25):
        linhas.append(
            {
                'tipo': 'Construtora',
                'obra': med_construtora.orcamento.obra.nome_obra,
                'planilha': med_construtora.orcamento.nome,
                'empreiteiro': '-',
                'numero': med_construtora.numero,
                'data_medicao': med_construtora.data_medicao,
                'periodo': f'{med_construtora.periodo_inicio:%d/%m/%Y} a {med_construtora.periodo_fim:%d/%m/%Y}',
                'medido': med_construtora.subtotal_periodo,
                'descontos': med_construtora.total_descontos,
                'liquido': med_construtora.total_liquido,
            }
        )
    for i in range(25):
        linhas.append(
            {
                'tipo': 'Empreiteiro Cumulativa',
                'obra': med_empreiteiro.obra.nome_obra,
                'planilha': med_empreiteiro.orcamento.nome,
                'empreiteiro': med_empreiteiro.empreiteiro,
                'numero': med_empreiteiro.numero,
                'data_medicao': med_empreiteiro.data_medicao,
                'periodo': f'{med_empreiteiro.periodo_inicio:%d/%m/%Y} a {med_empreiteiro.periodo_fim:%d/%m/%Y}',
                'medido': med_empreiteiro.subtotal_periodo,
                'descontos': med_empreiteiro.total_descontos,
                'liquido': med_empreiteiro.total_liquido,
            }
        )
    return linhas


def main():
    samples = []
    excels = {}
    runner = DiscoverRunner(verbosity=0, interactive=False)
    old_config = runner.setup_databases()
    try:
        with transaction.atomic():
            empresa = create_empresa('Empresa Demonstracao Engenharia', 'empresa-demo-qa', '#0f766e', '#e0f2fe')
            empresa_b = create_empresa('Empresa B Engenharia Teste', 'empresa-b-qa', '#7c2d12', '#ffedd5')
            obra, orcamento, med_construtora, med_empreiteiro = create_measurement_data(empresa)
            create_locacoes(empresa, obra)
            radar = create_radar(empresa)

            eventos, resumo = financial_events()
            col_fin = FinanceiroFiltroForm.COLUNAS_PADRAO
            col_med = RelatorioMedicoesForm.COLUNAS_PADRAO
            linhas_med = medicao_report_lines(med_construtora, med_empreiteiro)
            linhas_saldo, totais_saldo = _saldo_contratual_construtora(orcamento)

            pdf_outputs = {
                'financeiro.pdf': _financial_report_pdf(eventos, col_fin, resumo, empresa),
                'equipamentos_locados.pdf': relatorio_locacoes_equipamentos_pdf(make_request(empresa)).content,
                'radar_obras.pdf': _radar_obras_pdf(radar, {'arquivados': 'nao', 'ordenar': 'data_desc'}, empresa),
                'relatorio_medicoes.pdf': _pdf_relatorio_medicoes(linhas_med, col_med, _totais_relatorio_medicoes(linhas_med), empresa),
                'saldo_contratual.pdf': _pdf_saldo_contratual(orcamento, linhas_saldo, totais_saldo),
                'medicao_construtora.pdf': _pdf_medicao_construtora(med_construtora),
                'medicao_empreiteiro.pdf': _pdf_medicao_empreiteiro(med_empreiteiro),
                'empresa_b_relatorio_medicoes.pdf': _pdf_relatorio_medicoes(linhas_med[:8], col_med, _totais_relatorio_medicoes(linhas_med[:8]), empresa_b),
            }
            for name, content in pdf_outputs.items():
                path = write(OUT / name, content)
                samples.append((name, 'PDF', pdf_pages(path), str(path)))

            xlsx_outputs = {
                'financeiro.xlsx': _xlsx_relatorio_financeiro(eventos, col_fin, empresa),
                'relatorio_medicoes.xlsx': _xlsx_relatorio_medicoes(linhas_med, col_med, empresa),
                'saldo_contratual.xlsx': _xlsx_saldo_contratual(orcamento, linhas_saldo, totais_saldo),
                'medicao_construtora.xlsx': _xlsx_medicao(med_construtora, _itens_medicao_construtora_com_grupos(med_construtora)),
                'medicao_empreiteiro.xlsx': _xlsx_medicao(med_empreiteiro, med_empreiteiro.itens.select_related('item_orcamento')),
            }
            for name, content in xlsx_outputs.items():
                path = write(OUT / name, content)
                samples.append((name, 'XLSX', '-', str(path)))
                excels[name] = excel_flags(path)

            transaction.set_rollback(True)
    finally:
        runner.teardown_databases(old_config)

    print('SAMPLES')
    for row in samples:
        print('|'.join(map(str, row)))
    print('EXCEL')
    for name, flags in excels.items():
        print(name + '|' + '|'.join(f'{key}={value}' for key, value in flags.items()))


if __name__ == '__main__':
    main()
