import csv
import textwrap
import unicodedata
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from django.contrib import messages
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from controles.models import FaturamentoDireto
from controles.views import _build_simple_pdf
from documentos.excel import ExcelColumn, ExcelReportBuilder
from documentos.formatting import format_date_br, format_decimal_br, format_money_br, format_percent_br
from documentos.pdf import PdfDocument, PdfTableColumn, PdfTableGroup
from empresas.decorators import empresa_required
from empresas.documentos import draw_empresa_footer
from obras.models import Obra

from .forms import (
    EmpreiteiroForm,
    ImportarOrcamentoForm,
    ItemMedicaoConstrutoraFormSet,
    ItemMedicaoEmpreiteiroFormSet,
    ItemOrcamentoMedicaoFormSet,
    MedicaoConstrutoraCabecalhoForm,
    MedicaoConstrutoraForm,
    MedicaoEmpreiteiroCabecalhoForm,
    MedicaoEmpreiteiroForm,
    OrcamentoMedicaoManualForm,
    RelatorioMedicoesForm,
)
from .models import (
    Empreiteiro,
    FaturamentoDiretoMedicao,
    ItemMedicaoConstrutora,
    ItemMedicaoEmpreiteiro,
    ItemOrcamentoMedicao,
    MedicaoConstrutora,
    MedicaoEmpreiteiro,
    OrcamentoMedicao,
)


def _obras_empresa(empresa):
    return Obra.objects.filter(empresa=empresa)


def _orcamentos_empresa(empresa):
    return OrcamentoMedicao.objects.filter(obra__empresa=empresa)


def _medicoes_construtora_empresa(empresa):
    return MedicaoConstrutora.objects.filter(orcamento__obra__empresa=empresa)


def _medicoes_empreiteiro_empresa(empresa):
    return MedicaoEmpreiteiro.objects.filter(empresa=empresa)


def _money(value):
    return format_money_br(value)


def _clean_text(value):
    return unicodedata.normalize('NFKD', str(value or '-')).encode('ascii', 'ignore').decode('ascii')


def _font(size, bold=False):
    candidates = [
        settings.BASE_DIR / 'static' / 'fonts' / ('Arial Bold.ttf' if bold else 'Arial.ttf'),
        settings.BASE_DIR / 'static' / 'fonts' / ('arialbd.ttf' if bold else 'arial.ttf'),
        settings.BASE_DIR / 'static' / 'fonts' / ('DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf'),
        settings.BASE_DIR / 'static' / 'propostas' / 'fonts' / ('Arial Bold.ttf' if bold else 'Arial.ttf'),
        Path('C:/Windows/Fonts') / ('arialbd.ttf' if bold else 'arial.ttf'),
        Path('/usr/share/fonts/truetype/dejavu') / ('DejaVuSans-Bold.ttf' if bold else 'DejaVuSans.ttf'),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def _draw_wrapped_cell(draw, value, x, y, w, h, font, fill=(31, 41, 55), align='left'):
    text = _clean_text(value)
    avg_char_width = max(font.getlength('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz') / 52, 1)
    max_chars = max(int((w - 18) / avg_char_width), 8)
    lines = textwrap.wrap(text, width=max_chars) or ['']
    line_height = font.getbbox('Ag')[3] - font.getbbox('Ag')[1] + 7
    visible_lines = lines[: max(int((h - 12) / line_height), 1)]
    y_text = y + max((h - (line_height * len(visible_lines))) // 2, 6)
    for line in visible_lines:
        if align == 'right':
            x_text = x + w - font.getlength(line) - 10
        elif align == 'center':
            x_text = x + (w - font.getlength(line)) / 2
        else:
            x_text = x + 10
        draw.text((x_text, y_text), line, font=font, fill=fill)
        y_text += line_height


def _draw_pdf_table(draw, headers, rows, x, y, widths, row_h=54):
    border = (203, 213, 225)
    header_fill = (229, 236, 240)
    zebra = (248, 250, 252)
    header_font = _font(20, True)
    cell_font = _font(19)
    x_cursor = x
    for header, width in zip(headers, widths):
        draw.rectangle((x_cursor, y, x_cursor + width, y + row_h), fill=header_fill, outline=border, width=2)
        _draw_wrapped_cell(draw, header, x_cursor, y, width, row_h, header_font, fill=(15, 23, 42), align='center')
        x_cursor += width
    y += row_h
    for index, row in enumerate(rows):
        x_cursor = x
        fill = zebra if index % 2 else (255, 255, 255)
        for cell_index, (value, width) in enumerate(zip(row, widths)):
            draw.rectangle((x_cursor, y, x_cursor + width, y + row_h), fill=fill, outline=border, width=1)
            align = 'right' if cell_index >= len(row) - 3 else 'left'
            if cell_index in (0, 2):
                align = 'center'
            _draw_wrapped_cell(draw, value, x_cursor, y, width, row_h, cell_font, align=align)
            x_cursor += width
        y += row_h
    return y


def _draw_report_cell(
    draw,
    value,
    x,
    y,
    w,
    h,
    font,
    fill=(17, 24, 39),
    bg=None,
    border=(156, 163, 175),
    align='left',
    bold=False,
    width=1,
):
    if bg:
        draw.rectangle((x, y, x + w, y + h), fill=bg, outline=border, width=width)
    else:
        draw.rectangle((x, y, x + w, y + h), outline=border, width=width)
    text = _clean_text(value)
    avg_char_width = max(font.getlength('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz') / 52, 1)
    max_chars = max(int((w - 8) / avg_char_width), 4)
    lines = textwrap.wrap(text, width=max_chars) or ['']
    line_height = font.getbbox('Ag')[3] - font.getbbox('Ag')[1] + 4
    visible_lines = lines[: max(int((h - 6) / line_height), 1)]
    text_h = line_height * len(visible_lines)
    y_text = y + max((h - text_h) // 2, 3)
    for line in visible_lines:
        if align == 'right':
            x_text = x + w - font.getlength(line) - 5
        elif align == 'center':
            x_text = x + (w - font.getlength(line)) / 2
        else:
            x_text = x + 5
        draw.text((x_text, y_text), line, font=font, fill=fill)
        y_text += line_height


def _draw_report_row(draw, row, x, y, widths, height, font, bg=None, bold=False, cell_bgs=None):
    cursor = x
    for index, (value, width) in enumerate(zip(row, widths)):
        align = 'left' if index == 1 else 'center'
        cell_bg = cell_bgs[index] if cell_bgs and index < len(cell_bgs) and cell_bgs[index] else bg
        _draw_report_cell(draw, value, cursor, y, width, height, font, bg=cell_bg, align=align, width=1)
        cursor += width
    return y + height


def _fmt_qty(value):
    value = value or Decimal('0')
    quantized = value.quantize(Decimal('0.0001'))
    text = f'{quantized:,.4f}'.replace(',', 'X').replace('.', ',').replace('X', '.')
    return text.rstrip('0').rstrip(',')


def _percent_from_item(item):
    contrato = item.item_orcamento.quantidade or Decimal('0')
    if not contrato:
        return Decimal('0')
    percent = item.quantidade_acumulada_atual * Decimal('100') / contrato
    return percent.quantize(Decimal('0.01'))


def _pdf_medicao_construtora(medicao):
    itens = _itens_medicao_construtora_com_grupos(medicao)
    empresa = medicao.orcamento.obra.empresa
    doc = PdfDocument(
        empresa=empresa,
        title='Boletim de Medicoes',
        subtitle=f'Medicao no {medicao.numero}',
        orientation='landscape',
        filename=f'medicao_construtora_{medicao.numero}.pdf',
    )
    contract_bg = (232, 238, 247)
    measured_bg = (220, 245, 229)
    receivable_bg = (254, 226, 226)
    doc.add_title(emitted_on=date.today())
    doc.add_info_grid(
        [
            ('Contrato', getattr(medicao.orcamento, 'nome', '-')),
            ('Obra', medicao.orcamento.obra.nome_obra),
            ('Planilha', medicao.orcamento.nome),
            ('Cliente', medicao.orcamento.obra.cliente or '-'),
            ('Numero da medicao', medicao.numero),
            ('Periodo', f'{format_date_br(medicao.periodo_inicio)} a {format_date_br(medicao.periodo_fim)}'),
            ('Data da medicao', format_date_br(medicao.data_medicao)),
            ('Total do contrato', _money(medicao.orcamento.total_orcamento)),
            ('Total da obra', _money(medicao.orcamento.obra.contrato_atualizado)),
            ('Base da NF', _money(medicao.base_impostos)),
            ('Base INSS', _money(medicao.base_inss)),
            ('Total liquido', _money(medicao.total_liquido)),
        ],
        columns=4,
    )
    columns = [
        PdfTableColumn('ref', 'Ref.', width=58, align='center', bg=contract_bg),
        PdfTableColumn('descricao', 'Descricao', width=654, bg=contract_bg),
        PdfTableColumn('unidade', 'Un.', width=52, align='center', bg=contract_bg),
        PdfTableColumn('quantidade', 'Qtde', width=74, align='center', bg=contract_bg),
        PdfTableColumn('unit_material', 'Unit. material', width=125, align='right', bg=contract_bg),
        PdfTableColumn('unit_mao_obra', 'Unit. mao obra', width=125, align='right', bg=contract_bg),
        PdfTableColumn('unit_equip', 'Unit. equip.', width=118, align='right', bg=contract_bg),
        PdfTableColumn('preco_unit', 'Preco unit.', width=125, align='right', bg=contract_bg),
        PdfTableColumn('acumulado_anterior', 'Acum. anterior', width=94, align='center', bg=measured_bg),
        PdfTableColumn('medida', 'Medida', width=94, align='center', bg=measured_bg),
        PdfTableColumn('percentual', '%Exe.', width=76, align='center', bg=measured_bg),
        PdfTableColumn('material', 'Material', width=130, align='right', bg=receivable_bg),
        PdfTableColumn('mao_obra', 'Mao de obra', width=130, align='right', bg=receivable_bg),
        PdfTableColumn('equip', 'Equip.', width=124, align='right', bg=receivable_bg),
        PdfTableColumn('valor', 'Valor medicao', width=140, align='right', bg=receivable_bg),
    ]
    groups = [
        PdfTableGroup('Itens contratuais', 0, 8, contract_bg),
        PdfTableGroup('Itens medidos', 8, 11, measured_bg),
        PdfTableGroup('Valor a receber', 11, 15, receivable_bg),
    ]
    measured_cells = {'acumulado_anterior': measured_bg, 'medida': measured_bg, 'percentual': measured_bg}
    rows = []
    for item in itens:
        if isinstance(item, ItemOrcamentoMedicao) and item.eh_grupo:
            rows.append(
                {
                    'ref': item.item,
                    'descricao': item.descricao,
                    '__bg': doc.theme.surface,
                    '__bold': True,
                }
            )
            continue
        base = item.item_orcamento
        rows.append(
            {
                'ref': base.item,
                'descricao': base.descricao,
                'unidade': base.unidade or '-',
                'quantidade': _fmt_qty(base.quantidade),
                'unit_material': _money(base.preco_unitario_material),
                'unit_mao_obra': _money(base.preco_unitario_mao_obra),
                'unit_equip': _money(base.preco_unitario_equipamentos),
                'preco_unit': _money(base.preco_unitario_total),
                'acumulado_anterior': _fmt_qty(item.quantidade_acumulada_anterior),
                'medida': _fmt_qty(item.quantidade_periodo),
                'percentual': f'{_fmt_qty(_percent_from_item(item))}%',
                'material': _money(item.valor_material_periodo),
                'mao_obra': _money(item.valor_mao_obra_periodo),
                'equip': _money(item.valor_equipamentos_periodo),
                'valor': _money(item.valor_periodo),
                '__cell_bgs': measured_cells,
            }
        )
    for label, value, key in [
        ('Total material', medicao.total_material_periodo, 'material'),
        ('Total mao de obra', medicao.total_mao_obra_periodo, 'mao_obra'),
        ('Total equipamentos', medicao.total_equipamentos_periodo, 'equip'),
    ]:
        rows.append(
            {
                'descricao': label,
                key: _money(value),
                'valor': _money(value),
                '__bg': doc.theme.surface,
                '__bold': True,
            }
        )
    doc.add_table(columns, rows, row_height=38, groups=groups, table_body_level='small')

    desconto_adicional_nf = (
        medicao.desconto_adicional_calculado
        if medicao.desconto_adicional_reduz_base_nf
        else Decimal('0')
    )
    doc.add_totals_columns(
        [
            {
                'title': 'Total da medicao',
                'rows': [
                    ('Total medicao', _money(medicao.subtotal_periodo), False),
                    ('Desconto faturamento direto', f'- {_money(medicao.total_faturamento_direto)}', False),
                    ('Desconto adicional NF', f'- {_money(desconto_adicional_nf)}', False),
                    ('Total a faturar', _money(medicao.base_impostos), True),
                ],
            },
            {
                'title': 'Retencoes e impostos',
                'rows': [
                    ('Retencao tecnica', _money(medicao.retencao_tecnica_calculada), False),
                    ('INSS', _money(medicao.inss_calculado), False),
                    ('ISSQN', _money(medicao.issqn_calculado), False),
                    ('Total retido', _money(medicao.retencao_tecnica_calculada + medicao.inss_calculado + medicao.issqn_calculado), True),
                ],
            },
            {
                'title': 'Fechamento',
                'rows': [
                    ('Subtotal', _money(medicao.subtotal_periodo), False),
                    ('Base da NF', _money(medicao.base_impostos), False),
                    ('Material NF', _money(medicao.valor_material_nf), False),
                    ('Mao de obra NF', _money(medicao.valor_mao_obra_nf), False),
                    ('Equipamentos NF', _money(medicao.valor_equipamentos_nf), False),
                    ('Base INSS', _money(medicao.base_inss), False),
                    ('Descontos', _money(medicao.total_descontos), False),
                    ('Total liquido', _money(medicao.total_liquido), True),
                ],
            },
        ]
    )

    faturamentos = medicao.faturamentos_diretos.select_related('faturamento_direto')
    if faturamentos.exists():
        doc.add_section_header('Faturamentos diretos descontados', bg=doc.theme.header_fill)
        doc.add_table(
            [
                PdfTableColumn('numero', 'NF / OC', weight=0.8, align='center'),
                PdfTableColumn('empresa', 'Empresa que comprou', weight=1.4),
                PdfTableColumn('descricao', 'Descricao', weight=2.4),
                PdfTableColumn('percentual', '% descontado', weight=0.8, align='center'),
                PdfTableColumn('valor', 'Valor', weight=0.9, align='right'),
            ],
            [
                {
                    'numero': fd.faturamento_direto.numero_nf or fd.faturamento_direto.numero_ordem_compra or '-',
                    'empresa': fd.faturamento_direto.empresa_comprou,
                    'descricao': fd.faturamento_direto.descricao,
                    'percentual': format_percent_br(fd.percentual_descontado / Decimal('100')),
                    'valor': _money(fd.valor_descontado),
                }
                for fd in faturamentos
            ],
            row_height=42,
            table_body_level='small',
        )
    return doc.build()


def _decimal(value):
    if value in (None, ''):
        return Decimal('0')
    cleaned = str(value).strip().replace('R$', '').replace(' ', '')
    if ',' in cleaned:
        cleaned = cleaned.replace('.', '').replace(',', '.')
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return Decimal('0')


def _normalize_header(value):
    normalized = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    return ''.join(char for char in normalized.lower() if char.isalnum())


HEADER_ALIASES = {
    'item': ['item', 'codigo', 'cod'],
    'tipo': ['tipo', 'classe', 'categoria'],
    'descricao': ['descricao', 'descricaodoservico', 'servico'],
    'unidade': ['unidade', 'un', 'und'],
    'quantidade': ['quantidade', 'qtd', 'quant'],
    'preco_unitario_material': [
        'precounitariomaterial',
        'unitariomaterial',
        'material',
        'precounitario',
        'valorunitario',
        'unitario',
        'preco',
    ],
    'preco_unitario_mao_obra': ['precounitariomaodeobra', 'precounitariomaoobra', 'unitariomaodeobra', 'unitariomaoobra', 'maodeobra'],
    'preco_unitario_equipamentos': ['precounitarioequipamentos', 'unitarioequipamentos', 'equipamentos', 'equipamento'],
}


def _value(row, field):
    normalized = {_normalize_header(key): value for key, value in row.items()}
    for alias in HEADER_ALIASES[field]:
        if alias in normalized:
            return normalized[alias]
    return ''


def _is_integer_reference(value):
    text = str(value or '').strip()
    return bool(text) and text.isdigit()


def _tipo_item_orcamento(row, item, unidade, quantidade, material, mao_obra, equipamentos):
    tipo = _normalize_header(_value(row, 'tipo'))
    if tipo in {'grupo', 'titulo', 'disciplina', 'separador'}:
        return ItemOrcamentoMedicao.TIPO_GRUPO
    if tipo in {'item', 'servico', 'medivel', 'itemmedivel'}:
        return ItemOrcamentoMedicao.TIPO_ITEM
    if (
        _is_integer_reference(item)
        and not unidade
        and quantidade == Decimal('0')
        and material == Decimal('0')
        and mao_obra == Decimal('0')
        and equipamentos == Decimal('0')
    ):
        return ItemOrcamentoMedicao.TIPO_GRUPO
    return ItemOrcamentoMedicao.TIPO_ITEM


def _next_numero(model, **filters):
    last = model.objects.filter(**filters).order_by('-numero').first()
    return (last.numero + 1) if last else 1


def _percent_value(base, percent):
    if not percent:
        return None
    return (base * percent / Decimal('100')).quantize(Decimal('0.01'))


def _aplicar_percentuais_construtora(medicao):
    campos = {
        'retencao_tecnica': (medicao.subtotal_periodo, medicao.retencao_tecnica_percentual),
        'desconto_adicional': (medicao.subtotal_periodo, medicao.desconto_adicional_percentual),
        'issqn': (medicao.base_impostos, medicao.issqn_percentual),
        'inss': (medicao.base_inss, medicao.inss_percentual),
    }
    updates = []
    for field, (base, percent) in campos.items():
        value = _percent_value(base, percent)
        if value is not None:
            setattr(medicao, field, value)
            updates.append(field)
    if updates:
        medicao.save(update_fields=updates + ['updated_at'])


def _atualizar_resumo_faturamento_direto(faturamento):
    vinculos = faturamento.vinculos_medicao.select_related('medicao').order_by('medicao__numero', 'id')
    partes = [f'{vinculo.medicao.label_medicao} ({vinculo.percentual_descontado:.2f}%)' for vinculo in vinculos]
    faturamento.medicao_desconto = ' / '.join(partes)[:120]
    faturamento.save(update_fields=['medicao_desconto', 'updated_at'])


def _sync_faturamentos_diretos(medicao, post_data):
    atuais = {
        vinculo.faturamento_direto_id: vinculo
        for vinculo in medicao.faturamentos_diretos.select_related('faturamento_direto')
    }
    usados = set()
    for faturamento in FaturamentoDireto.objects.filter(obra=medicao.orcamento.obra):
        raw_percent = (post_data.get(f'faturamento_direto_{faturamento.id}_percentual') or '').strip()
        try:
            percentual = Decimal(raw_percent.replace(',', '.')) if raw_percent else Decimal('0')
        except InvalidOperation:
            percentual = Decimal('0')
        percentual = max(Decimal('0'), min(percentual, Decimal('100')))
        ja_descontado = sum(
            (
                vinculo.percentual_descontado
                for vinculo in faturamento.vinculos_medicao.exclude(medicao=medicao)
            ),
            Decimal('0'),
        )
        saldo_percentual = max(Decimal('100') - ja_descontado, Decimal('0'))
        percentual = min(percentual, saldo_percentual)
        vinculo = atuais.get(faturamento.id)
        if percentual > 0:
            if not vinculo:
                vinculo = FaturamentoDiretoMedicao(medicao=medicao, faturamento_direto=faturamento)
            vinculo.percentual_descontado = percentual
            vinculo.save()
            usados.add(faturamento.id)
        elif vinculo:
            vinculo.delete()
        _atualizar_resumo_faturamento_direto(faturamento)
    for faturamento_id, vinculo in atuais.items():
        if faturamento_id not in usados and not FaturamentoDireto.objects.filter(
            id=faturamento_id,
            obra=medicao.orcamento.obra,
        ).exists():
            faturamento = vinculo.faturamento_direto
            vinculo.delete()
            _atualizar_resumo_faturamento_direto(faturamento)


def _faturamentos_diretos_context(medicao):
    linhas = []
    for faturamento in FaturamentoDireto.objects.filter(obra=medicao.orcamento.obra).order_by('data_lancamento', 'id'):
        vinculo_atual = medicao.faturamentos_diretos.filter(faturamento_direto=faturamento).first()
        percentual_atual = vinculo_atual.percentual_descontado if vinculo_atual else Decimal('0')
        percentual_outros = sum(
            (
                vinculo.percentual_descontado
                for vinculo in faturamento.vinculos_medicao.exclude(medicao=medicao)
            ),
            Decimal('0'),
        )
        saldo_percentual = max(Decimal('100') - percentual_outros, Decimal('0'))
        if saldo_percentual <= 0 and not vinculo_atual:
            continue
        linhas.append(
            {
                'faturamento': faturamento,
                'percentual_atual': percentual_atual,
                'percentual_outros': percentual_outros,
                'saldo_percentual': saldo_percentual,
                'valor_atual': vinculo_atual.valor_descontado if vinculo_atual else Decimal('0'),
            }
        )
    return linhas


def _aplicar_percentuais_empreiteiro(medicao):
    base = medicao.subtotal_periodo
    campos = {
        'retencao_tecnica': medicao.retencao_tecnica_percentual,
        'desconto_adicional': medicao.desconto_adicional_percentual,
    }
    updates = []
    for field, percent in campos.items():
        value = _percent_value(base, percent)
        if value is not None:
            setattr(medicao, field, value)
            updates.append(field)
    if updates:
        medicao.save(update_fields=updates + ['updated_at'])


def _empreiteiros_json(empresa):
    return [
        {
            'id': empreiteiro.id,
            'nome': empreiteiro.nome,
            'cpf_cnpj': empreiteiro.cpf_cnpj,
            'pix': empreiteiro.pix,
        }
        for empreiteiro in Empreiteiro.objects.filter(empresa=empresa, ativo=True).order_by('nome')
    ]


def _sync_empreiteiro_medicao(medicao):
    if medicao.empreiteiro_cadastro_id:
        cadastro = medicao.empreiteiro_cadastro
        medicao.empreiteiro = cadastro.nome
        medicao.cpf_cnpj = cadastro.cpf_cnpj
        medicao.pix = cadastro.pix
        medicao.save(update_fields=['empreiteiro', 'cpf_cnpj', 'pix', 'updated_at'])
        return cadastro
    return None


def _read_csv(file):
    raw = file.read()
    if not raw:
        return None, 'O arquivo CSV esta vazio.'
    for encoding in ['utf-8-sig', 'latin-1']:
        try:
            content = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        content = raw.decode('utf-8', errors='ignore')
    sample = content[:2048]
    delimiter = ';' if sample.count(';') >= sample.count(',') else ','
    reader = csv.DictReader(StringIO(content), delimiter=delimiter)
    if not reader.fieldnames:
        return None, 'O CSV precisa ter cabecalho na primeira linha.'
    normalized_headers = {_normalize_header(header) for header in reader.fieldnames}
    if not any(alias in normalized_headers for alias in HEADER_ALIASES['descricao']):
        return None, 'Nao encontrei a coluna descricao no CSV.'
    return reader, ''


def medicoes_home(request):
    return redirect('medicoes_construtora_home')


RELATORIO_MEDICOES_COLUNAS = dict(RelatorioMedicoesForm.COLUNAS_CHOICES)


def _format_date_br(value):
    return value.strftime('%d/%m/%Y') if value else '-'


def _format_periodo(inicio, fim):
    return f'{_format_date_br(inicio)} a {_format_date_br(fim)}'


def _format_percent(value):
    if value in (None, ''):
        return '-'
    return f'{value:.2f}%'.replace('.', ',')


def _relatorio_linha_display(row, coluna):
    if coluna in {'medido', 'descontos', 'liquido'}:
        return _money(row[coluna])
    if coluna == 'data_medicao':
        return _format_date_br(row[coluna])
    if coluna == 'periodo':
        return _format_periodo(row['periodo_inicio'], row['periodo_fim'])
    if coluna == 'percentual':
        return _format_percent(row[coluna])
    return row.get(coluna) or '-'


def _linhas_relatorio_medicoes(filtros, empresa):
    tipo = filtros.get('tipo') or ''
    obra = filtros.get('obra')
    empreiteiro = filtros.get('empreiteiro')
    data_inicial = filtros.get('data_inicial')
    data_final = filtros.get('data_final')
    linhas = []

    if tipo in {'', 'construtora'} and not empreiteiro:
        medicoes = _medicoes_construtora_empresa(empresa).select_related('orcamento', 'orcamento__obra')
        if obra:
            medicoes = medicoes.filter(orcamento__obra=obra)
        if data_inicial:
            medicoes = medicoes.filter(data_medicao__gte=data_inicial)
        if data_final:
            medicoes = medicoes.filter(data_medicao__lte=data_final)
        for medicao in medicoes:
            linhas.append(
                {
                    'tipo': 'Construtora',
                    'obra': medicao.orcamento.obra.nome_obra,
                    'planilha': medicao.orcamento.nome,
                    'empreiteiro': '-',
                    'numero': medicao.numero,
                    'data_medicao': medicao.data_medicao,
                    'periodo_inicio': medicao.periodo_inicio,
                    'periodo_fim': medicao.periodo_fim,
                    'medido': medicao.subtotal_periodo,
                    'descontos': medicao.total_descontos,
                    'liquido': medicao.total_liquido,
                    'percentual': medicao.orcamento.percentual_medido_construtora,
                    'url': reverse('editar_medicao_construtora', args=[medicao.id]),
                }
            )

    if tipo in {'', 'empreiteiro', 'empreiteiro_simples', 'empreiteiro_cumulativa'}:
        medicoes = _medicoes_empreiteiro_empresa(empresa).select_related('obra', 'orcamento', 'orcamento__obra', 'empreiteiro_cadastro')
        if tipo == 'empreiteiro_simples':
            medicoes = medicoes.filter(tipo=MedicaoEmpreiteiro.TIPO_SIMPLES)
        elif tipo == 'empreiteiro_cumulativa':
            medicoes = medicoes.filter(tipo=MedicaoEmpreiteiro.TIPO_CUMULATIVA)
        if obra:
            medicoes = medicoes.filter(Q(obra=obra) | Q(orcamento__obra=obra))
        if empreiteiro:
            medicoes = medicoes.filter(empreiteiro_cadastro=empreiteiro)
        if data_inicial:
            medicoes = medicoes.filter(data_medicao__gte=data_inicial)
        if data_final:
            medicoes = medicoes.filter(data_medicao__lte=data_final)
        for medicao in medicoes:
            orcamento = medicao.orcamento
            linhas.append(
                {
                    'tipo': f'Empreiteiro {medicao.get_tipo_display()}',
                    'obra': str(medicao.obra or getattr(orcamento, 'obra', '-') or '-'),
                    'planilha': getattr(orcamento, 'nome', '-') or '-',
                    'empreiteiro': medicao.empreiteiro,
                    'numero': medicao.numero,
                    'data_medicao': medicao.data_medicao,
                    'periodo_inicio': medicao.periodo_inicio,
                    'periodo_fim': medicao.periodo_fim,
                    'medido': medicao.subtotal_periodo,
                    'descontos': medicao.total_descontos,
                    'liquido': medicao.total_liquido,
                    'percentual': orcamento.percentual_medido_empreiteiro if orcamento else None,
                    'url': reverse('editar_medicao_empreiteiro', args=[medicao.id]),
                }
            )

    return sorted(linhas, key=lambda row: (row['data_medicao'], row['tipo'], str(row['obra'])), reverse=True)


def _totais_relatorio_medicoes(linhas):
    return {
        'medido': sum((linha['medido'] for linha in linhas), Decimal('0')),
        'descontos': sum((linha['descontos'] for linha in linhas), Decimal('0')),
        'liquido': sum((linha['liquido'] for linha in linhas), Decimal('0')),
    }


def _xlsx_relatorio_medicoes(linhas, colunas, empresa):
    excel_columns = []
    for coluna in colunas:
        width = 16
        align = 'center'
        number_format = None
        wrap = False
        if coluna in {'obra', 'planilha', 'empreiteiro'}:
            width = 34
            align = 'left'
            wrap = True
        elif coluna in {'medido', 'descontos', 'liquido'}:
            width = 18
            align = 'right'
            number_format = ExcelReportBuilder.MONEY_FORMAT
        elif coluna == 'data_medicao':
            width = 15
            number_format = ExcelReportBuilder.DATE_FORMAT
        elif coluna == 'percentual':
            width = 14
            align = 'right'
            number_format = ExcelReportBuilder.PERCENT_FORMAT
        excel_columns.append(
            ExcelColumn(
                key=coluna,
                label=RELATORIO_MEDICOES_COLUNAS[coluna],
                width=width,
                align=align,
                number_format=number_format,
                wrap=wrap,
            )
        )
    rows = []
    for linha in linhas:
        row = {}
        for coluna in colunas:
            if coluna in {'medido', 'descontos', 'liquido'}:
                row[coluna] = linha[coluna]
            elif coluna == 'data_medicao':
                row[coluna] = linha[coluna]
            elif coluna == 'percentual':
                row[coluna] = (linha[coluna] or Decimal('0')) / Decimal('100')
            elif coluna == 'periodo':
                row[coluna] = _format_periodo(linha['periodo_inicio'], linha['periodo_fim'])
            else:
                row[coluna] = linha.get(coluna) or '-'
        rows.append(row)
    builder = ExcelReportBuilder(empresa=empresa, title='Relatorio gerencial de medicoes', sheet_name='Medicoes')
    builder.add_header(emitted_on=date.today())
    builder.add_table(excel_columns, rows)
    return builder.build()


def _pdf_relatorio_medicoes(linhas, colunas, totais, empresa):
    colunas = colunas or RelatorioMedicoesForm.COLUNAS_PADRAO
    weights = {
        'tipo': 1.2,
        'obra': 1.8,
        'planilha': 1.8,
        'empreiteiro': 1.8,
        'numero': 0.75,
        'data_medicao': 1.0,
        'periodo': 1.3,
        'medido': 1.05,
        'descontos': 1.05,
        'liquido': 1.05,
        'percentual': 0.9,
    }
    pdf_columns = []
    for coluna in colunas:
        align = 'center'
        if coluna in {'obra', 'planilha', 'empreiteiro'}:
            align = 'left'
        elif coluna in {'medido', 'descontos', 'liquido', 'percentual'}:
            align = 'right'
        pdf_columns.append(PdfTableColumn(coluna, RELATORIO_MEDICOES_COLUNAS[coluna], weight=weights.get(coluna, 1), align=align))
    rows = [
        {coluna: _relatorio_linha_display(linha, coluna) for coluna in colunas}
        for linha in linhas
    ]
    doc = PdfDocument(
        empresa=empresa,
        title='Relatorio gerencial de medicoes',
        subtitle=f'{len(linhas)} registro(s)',
        orientation='landscape',
        filename='relatorio_medicoes.pdf',
    )
    doc.add_title(emitted_on=date.today())
    doc.add_info_grid(
        [
            ('Total medido', _money(totais['medido'])),
            ('Descontos', _money(totais['descontos'])),
            ('Total liquido', _money(totais['liquido'])),
        ],
        columns=3,
    )
    doc.add_table(pdf_columns, rows, row_height=44)
    return doc.build()


def relatorio_medicoes(request):
    data = request.GET.copy()
    if 'colunas' not in data:
        data.setlist('colunas', RelatorioMedicoesForm.COLUNAS_PADRAO)
    form = RelatorioMedicoesForm(data, empresa=request.empresa)
    linhas = []
    totais = _totais_relatorio_medicoes(linhas)
    colunas = RelatorioMedicoesForm.COLUNAS_PADRAO
    if form.is_valid():
        colunas = form.cleaned_data['colunas']
        linhas = _linhas_relatorio_medicoes(form.cleaned_data, request.empresa)
        totais = _totais_relatorio_medicoes(linhas)
        export = request.GET.get('export')
        if export == 'excel':
            response = HttpResponse(
                _xlsx_relatorio_medicoes(linhas, colunas, request.empresa),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="relatorio_medicoes.xlsx"'
            return response
        if export == 'pdf':
            response = HttpResponse(_pdf_relatorio_medicoes(linhas, colunas, totais, request.empresa), content_type='application/pdf')
            response['Content-Disposition'] = 'inline; filename="relatorio_medicoes.pdf"'
            return response

    query_params = request.GET.copy()
    query_params.pop('export', None)
    query_string = query_params.urlencode()
    return render(
        request,
        'medicoes/relatorio_medicoes.html',
        {
            'form': form,
            'linhas': linhas,
            'totais': totais,
            'colunas': colunas,
            'colunas_labels': RELATORIO_MEDICOES_COLUNAS,
            'query_string': query_string,
        },
    )


def medicoes_construtora_home(request):
    contexto = {
        'obras': _obras_empresa(request.empresa).filter(
            orcamentos_medicao__tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
        ).distinct().order_by('nome_obra'),
        'planilhas': _orcamentos_empresa(request.empresa).filter(
            tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
        ).select_related('obra').prefetch_related('medicoes_construtora', 'itens'),
        'medicoes': _medicoes_construtora_empresa(request.empresa).select_related('orcamento', 'orcamento__obra')[:12],
    }
    return render(request, 'medicoes/construtora_home.html', contexto)


def medicoes_empreiteiros_home(request):
    contexto = {
        'simples': _medicoes_empreiteiro_empresa(request.empresa).filter(
            tipo=MedicaoEmpreiteiro.TIPO_SIMPLES,
        ).select_related('obra')[:15],
        'cumulativas': _medicoes_empreiteiro_empresa(request.empresa).filter(
            tipo=MedicaoEmpreiteiro.TIPO_CUMULATIVA,
        ).select_related('obra', 'orcamento')[:15],
        'planilhas': _orcamentos_empresa(request.empresa).filter(
            tipo=OrcamentoMedicao.TIPO_EMPREITEIRO,
        ).select_related('obra').prefetch_related('itens', 'medicoes_empreiteiro')[:15],
    }
    return render(request, 'medicoes/empreiteiros_home.html', contexto)


def lista_empreiteiros(request):
    empreiteiros = Empreiteiro.objects.filter(empresa=request.empresa)
    busca = request.GET.get('busca', '').strip()
    if busca:
        empreiteiros = empreiteiros.filter(
            Q(nome__icontains=busca)
            | Q(cpf_cnpj__icontains=busca)
            | Q(pix__icontains=busca)
        )
    return render(
        request,
        'medicoes/lista_empreiteiros.html',
        {'empreiteiros': empreiteiros, 'busca': busca},
    )


def novo_empreiteiro(request):
    if request.method == 'POST':
        form = EmpreiteiroForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empreiteiro cadastrado com sucesso.')
            return redirect('lista_empreiteiros_medicao')
    else:
        form = EmpreiteiroForm(empresa=request.empresa)
    return render(request, 'medicoes/form_empreiteiro.html', {'form': form, 'titulo': 'Novo empreiteiro'})


def editar_empreiteiro(request, empreiteiro_id):
    empreiteiro = get_object_or_404(Empreiteiro, id=empreiteiro_id, empresa=request.empresa)
    if request.method == 'POST':
        form = EmpreiteiroForm(request.POST, instance=empreiteiro, empresa=request.empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empreiteiro atualizado com sucesso.')
            return redirect('lista_empreiteiros_medicao')
    else:
        form = EmpreiteiroForm(instance=empreiteiro, empresa=request.empresa)
    return render(
        request,
        'medicoes/form_empreiteiro.html',
        {'form': form, 'titulo': 'Editar empreiteiro', 'empreiteiro': empreiteiro},
    )


def medicoes_obra(request, obra_id):
    obra = get_object_or_404(Obra, id=obra_id, empresa=request.empresa)
    planilhas = obra.orcamentos_medicao.prefetch_related('itens', 'medicoes_construtora', 'medicoes_empreiteiro')
    medicoes_construtora = MedicaoConstrutora.objects.filter(orcamento__obra=obra).select_related('orcamento')
    medicoes_empreiteiro = _medicoes_empreiteiro_empresa(request.empresa).filter(obra=obra).select_related('orcamento')
    return render(
        request,
        'medicoes/obra.html',
        {
            'obra': obra,
            'planilhas': planilhas,
            'medicoes_construtora': medicoes_construtora,
            'medicoes_empreiteiro': medicoes_empreiteiro,
        },
    )


def lista_orcamentos(request):
    orcamentos = _orcamentos_empresa(request.empresa).select_related('obra')
    return render(request, 'medicoes/lista_orcamentos.html', {'orcamentos': orcamentos})


def importar_orcamento(request):
    if request.method == 'POST':
        form = ImportarOrcamentoForm(request.POST, request.FILES, empresa=request.empresa)
        if form.is_valid():
            reader, error = _read_csv(form.cleaned_data['arquivo'])
            if error:
                form.add_error('arquivo', error)
                return render(request, 'medicoes/importar_orcamento.html', {'form': form})
            with transaction.atomic():
                orcamento = OrcamentoMedicao.objects.create(
                    obra=form.cleaned_data['obra'],
                    nome=form.cleaned_data['nome'],
                    tipo=form.cleaned_data['tipo'],
                    observacoes=form.cleaned_data['observacoes'],
                )
                itens = []
                for row in reader:
                    descricao = _value(row, 'descricao').strip()
                    if not descricao:
                        continue
                    item_ref = _value(row, 'item').strip() or str(len(itens) + 1)
                    unidade = _value(row, 'unidade').strip()
                    quantidade = _decimal(_value(row, 'quantidade'))
                    material = _decimal(_value(row, 'preco_unitario_material'))
                    mao_obra = _decimal(_value(row, 'preco_unitario_mao_obra'))
                    equipamentos = _decimal(_value(row, 'preco_unitario_equipamentos'))
                    tipo_item = _tipo_item_orcamento(row, item_ref, unidade, quantidade, material, mao_obra, equipamentos)
                    if tipo_item == ItemOrcamentoMedicao.TIPO_GRUPO:
                        unidade = ''
                        quantidade = Decimal('0')
                        material = Decimal('0')
                        mao_obra = Decimal('0')
                        equipamentos = Decimal('0')
                    itens.append(
                        ItemOrcamentoMedicao(
                            orcamento=orcamento,
                            tipo=tipo_item,
                            ordem=len(itens),
                            item=item_ref,
                            descricao=descricao,
                            unidade=unidade,
                            quantidade=quantidade,
                            preco_unitario_material=material,
                            preco_unitario_mao_obra=mao_obra,
                            preco_unitario_equipamentos=equipamentos,
                        )
                    )
                ItemOrcamentoMedicao.objects.bulk_create(itens)
                if not itens:
                    transaction.set_rollback(True)
                    form.add_error('arquivo', 'Nenhum item valido foi encontrado no CSV.')
                    return render(request, 'medicoes/importar_orcamento.html', {'form': form})
            messages.success(request, f'Planilha de medicao importada com {len(itens)} itens.')
            return redirect('detalhe_orcamento_medicao', orcamento_id=orcamento.id)
    else:
        initial = {}
        if request.GET.get('obra'):
            initial['obra'] = request.GET['obra']
        form = ImportarOrcamentoForm(initial=initial, empresa=request.empresa)
    return render(request, 'medicoes/importar_orcamento.html', {'form': form})


def novo_orcamento_manual(request):
    tipo = request.GET.get('tipo')
    if tipo not in {OrcamentoMedicao.TIPO_CONSTRUTORA, OrcamentoMedicao.TIPO_EMPREITEIRO}:
        tipo = OrcamentoMedicao.TIPO_CONSTRUTORA
    initial = {'tipo': tipo}
    if request.GET.get('obra'):
        initial['obra'] = request.GET['obra']
    if request.method == 'POST':
        form = OrcamentoMedicaoManualForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            orcamento = form.save()
            messages.success(request, 'Planilha manual criada. Agora adicione os itens para usar nas medicoes.')
            return redirect('editar_itens_orcamento_medicao', orcamento_id=orcamento.id)
    else:
        form = OrcamentoMedicaoManualForm(initial=initial, empresa=request.empresa)
    return render(
        request,
        'medicoes/form_orcamento_manual.html',
        {
            'form': form,
            'titulo': 'Nova planilha manual de medicao',
        },
    )


def detalhe_orcamento(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra').prefetch_related(
            'itens',
            'medicoes_construtora',
            'medicoes_empreiteiro',
        ),
        id=orcamento_id,
        obra__empresa=request.empresa,
    )
    itens = orcamento.itens.all()
    medicoes_construtora = orcamento.medicoes_construtora.all()
    medicoes_empreiteiro = orcamento.medicoes_empreiteiro.all()
    return render(
        request,
        'medicoes/detalhe_orcamento.html',
        {
            'orcamento': orcamento,
            'itens': itens,
            'medicoes_construtora': medicoes_construtora,
            'medicoes_empreiteiro': medicoes_empreiteiro,
        },
    )


def _saldo_contratual_construtora(orcamento):
    medidos = {
        row['item_orcamento_id']: row['total'] or Decimal('0')
        for row in ItemMedicaoConstrutora.objects.filter(medicao__orcamento=orcamento)
        .values('item_orcamento_id')
        .annotate(total=Sum('quantidade_periodo'))
    }
    linhas = []
    totais = {
        'material': Decimal('0'),
        'mao_obra': Decimal('0'),
        'equipamentos': Decimal('0'),
        'total': Decimal('0'),
    }

    for item in orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_ITEM):
        quantidade_medida = medidos.get(item.id, Decimal('0'))
        saldo_quantidade = max(item.quantidade - quantidade_medida, Decimal('0'))
        if saldo_quantidade <= 0:
            continue

        saldo_material = saldo_quantidade * item.preco_unitario_material
        saldo_mao_obra = saldo_quantidade * item.preco_unitario_mao_obra
        saldo_equipamentos = saldo_quantidade * item.preco_unitario_equipamentos
        saldo_total = saldo_material + saldo_mao_obra + saldo_equipamentos
        linhas.append(
            {
                'item': item,
                'quantidade_medida': quantidade_medida,
                'saldo_quantidade': saldo_quantidade,
                'saldo_material': saldo_material,
                'saldo_mao_obra': saldo_mao_obra,
                'saldo_equipamentos': saldo_equipamentos,
                'saldo_total': saldo_total,
            }
        )
        totais['material'] += saldo_material
        totais['mao_obra'] += saldo_mao_obra
        totais['equipamentos'] += saldo_equipamentos
        totais['total'] += saldo_total

    return linhas, totais


def _linhas_medicao_construtora_formset(medicao, formset):
    forms_by_item = {form.instance.item_orcamento_id: form for form in formset.forms}
    linhas = []
    for item in medicao.orcamento.itens.all():
        if item.eh_grupo:
            linhas.append({'tipo': 'grupo', 'item': item})
        elif item.id in forms_by_item:
            linhas.append({'tipo': 'item', 'form': forms_by_item[item.id]})
    return linhas


def _sincronizar_itens_medicao_construtora(medicao):
    medicao.itens.filter(item_orcamento__tipo=ItemOrcamentoMedicao.TIPO_GRUPO).delete()
    itens_existentes = set(medicao.itens.values_list('item_orcamento_id', flat=True))
    itens_novos = [
        ItemMedicaoConstrutora(medicao=medicao, item_orcamento=item)
        for item in medicao.orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_ITEM)
        if item.id not in itens_existentes
    ]
    if itens_novos:
        ItemMedicaoConstrutora.objects.bulk_create(itens_novos)


def _itens_medicao_construtora_com_grupos(medicao):
    itens_medicao = {
        item.item_orcamento_id: item
        for item in medicao.itens.select_related('item_orcamento')
    }
    linhas = []
    for item_orcamento in medicao.orcamento.itens.all():
        if item_orcamento.eh_grupo:
            linhas.append(item_orcamento)
        elif item_orcamento.id in itens_medicao:
            linhas.append(itens_medicao[item_orcamento.id])
    return linhas


def saldo_contratual_construtora(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra').prefetch_related('itens'),
        id=orcamento_id,
        obra__empresa=request.empresa,
        tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
    )
    linhas, totais = _saldo_contratual_construtora(orcamento)
    return render(
        request,
        'medicoes/saldo_contratual_construtora.html',
        {
            'orcamento': orcamento,
            'linhas': linhas,
            'totais': totais,
        },
    )


def saldo_contratual_construtora_pdf(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra').prefetch_related('itens'),
        id=orcamento_id,
        obra__empresa=request.empresa,
        tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
    )
    linhas, totais = _saldo_contratual_construtora(orcamento)
    response = HttpResponse(_pdf_saldo_contratual(orcamento, linhas, totais), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="saldo_contratual_{orcamento.id}.pdf"'
    return response


def saldo_contratual_construtora_excel(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra').prefetch_related('itens'),
        id=orcamento_id,
        obra__empresa=request.empresa,
        tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
    )
    linhas, totais = _saldo_contratual_construtora(orcamento)
    response = HttpResponse(
        _xlsx_saldo_contratual(orcamento, linhas, totais),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="saldo_contratual_{orcamento.id}.xlsx"'
    return response


def _normalizar_ordem_itens_orcamento(orcamento):
    itens = list(orcamento.itens.order_by('ordem', 'id'))
    for index, item in enumerate(itens):
        if item.ordem != index:
            item.ordem = index
            item.save(update_fields=['ordem'])


def _xlsx_saldo_contratual(orcamento, linhas, totais):
    columns = [
        ExcelColumn('ref', 'Ref.', width=10, align='center'),
        ExcelColumn('descricao', 'Descricao', width=58, align='left', wrap=True),
        ExcelColumn('unidade', 'Un.', width=8, align='center'),
        ExcelColumn('quantidade', 'Qtde contrato', width=14, align='right', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
        ExcelColumn('medido', 'Acum. medido', width=14, align='right', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
        ExcelColumn('saldo_quantidade', 'Saldo', width=14, align='right', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
        ExcelColumn('unit_material', 'Unit. material', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('unit_mao_obra', 'Unit. mao obra', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('unit_equip', 'Unit. equip.', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('saldo_material', 'Material saldo', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('saldo_mao_obra', 'Mao obra saldo', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('saldo_equip', 'Equip. saldo', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ExcelColumn('saldo_total', 'Total saldo', width=17, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
    ]
    rows = []
    for linha in linhas:
        item = linha['item']
        rows.append(
            {
                'ref': item.item,
                'descricao': item.descricao,
                'unidade': item.unidade,
                'quantidade': item.quantidade,
                'medido': linha['quantidade_medida'],
                'saldo_quantidade': linha['saldo_quantidade'],
                'unit_material': item.preco_unitario_material,
                'unit_mao_obra': item.preco_unitario_mao_obra,
                'unit_equip': item.preco_unitario_equipamentos,
                'saldo_material': linha['saldo_material'],
                'saldo_mao_obra': linha['saldo_mao_obra'],
                'saldo_equip': linha['saldo_equipamentos'],
                'saldo_total': linha['saldo_total'],
            }
        )
    builder = ExcelReportBuilder(
        empresa=orcamento.obra.empresa,
        title='Saldo contratual da construtora',
        subtitle=f'Obra: {orcamento.obra} | Planilha: {orcamento.nome}',
        sheet_name='Saldo contratual',
    )
    builder.add_header(emitted_on=date.today())
    builder.add_table(columns, rows)
    ws = builder.ws
    total_row = builder.current_row
    ws.cell(total_row, 9, 'Totais')
    total_values = [totais['material'], totais['mao_obra'], totais['equipamentos'], totais['total']]
    for offset, value in enumerate(total_values, start=10):
        cell = ws.cell(total_row, offset, value)
        cell.number_format = ExcelReportBuilder.MONEY_FORMAT
        cell.alignment = Alignment(horizontal='right', vertical='center')
    for col in range(9, 14):
        cell = ws.cell(total_row, col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='F3F4F6')
        cell.border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1'),
        )
    return builder.build()


def _pdf_saldo_contratual(orcamento, linhas, totais):
    doc = PdfDocument(
        empresa=orcamento.obra.empresa,
        title='Saldo contratual da construtora',
        subtitle=f'Obra: {orcamento.obra} | Planilha: {orcamento.nome}',
        orientation='landscape',
        filename=f'saldo_contratual_{orcamento.id}.pdf',
    )
    doc.add_title(emitted_on=date.today())
    doc.add_info_grid(
        [
            ('Obra', orcamento.obra),
            ('Planilha', orcamento.nome),
            ('Total contrato', _money(orcamento.total_orcamento)),
            ('Emitido em', format_date_br(date.today())),
        ],
        columns=4,
    )
    rows = []
    for linha in linhas:
        item = linha['item']
        rows.append(
            {
                'ref': item.item,
                'descricao': item.descricao,
                'unidade': item.unidade or '-',
                'quantidade': format_decimal_br(item.quantidade, 4),
                'medido': format_decimal_br(linha['quantidade_medida'], 4),
                'saldo': format_decimal_br(linha['saldo_quantidade'], 4),
                'unit_total': _money(item.preco_unitario_total),
                'material': _money(linha['saldo_material']),
                'mao_obra': _money(linha['saldo_mao_obra']),
                'equip': _money(linha['saldo_equipamentos']),
                'total': _money(linha['saldo_total']),
            }
        )
    doc.add_table(
        [
            PdfTableColumn('ref', 'Ref.', weight=0.6, align='center'),
            PdfTableColumn('descricao', 'Descricao', weight=3.7),
            PdfTableColumn('unidade', 'Un.', weight=0.55, align='center'),
            PdfTableColumn('quantidade', 'Contrato', weight=0.9, align='right'),
            PdfTableColumn('medido', 'Medido', weight=0.9, align='right'),
            PdfTableColumn('saldo', 'Saldo', weight=0.9, align='right'),
            PdfTableColumn('unit_total', 'Unit. total', weight=1.1, align='right'),
            PdfTableColumn('material', 'Material', weight=1.2, align='right'),
            PdfTableColumn('mao_obra', 'Mao obra', weight=1.2, align='right'),
            PdfTableColumn('equip', 'Equip.', weight=1.1, align='right'),
            PdfTableColumn('total', 'Total saldo', weight=1.45, align='right'),
        ],
        rows,
        row_height=44,
    )
    doc.add_totals_box(
        [
            ('Material a medir', _money(totais['material']), False),
            ('Mao de obra a medir', _money(totais['mao_obra']), False),
            ('Equipamentos a medir', _money(totais['equipamentos']), False),
            ('Total ainda em contrato', _money(totais['total']), True),
        ],
        width=760,
    )
    return doc.build()


def editar_itens_orcamento(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra'),
        id=orcamento_id,
        obra__empresa=request.empresa,
    )
    if request.method == 'POST':
        formset = ItemOrcamentoMedicaoFormSet(request.POST, instance=orcamento)
        if formset.is_valid():
            formset.save()
            _normalizar_ordem_itens_orcamento(orcamento)
            messages.success(request, 'Itens da planilha atualizados com sucesso.')
            return redirect('detalhe_orcamento_medicao', orcamento_id=orcamento.id)
    else:
        formset = ItemOrcamentoMedicaoFormSet(instance=orcamento)
    return render(
        request,
        'medicoes/editar_itens_orcamento.html',
        {
            'orcamento': orcamento,
            'formset': formset,
        },
    )


def excluir_orcamento(request, orcamento_id):
    orcamento = get_object_or_404(
        OrcamentoMedicao.objects.select_related('obra'),
        id=orcamento_id,
        obra__empresa=request.empresa,
    )
    obra_id = orcamento.obra_id
    if request.method == 'POST':
        orcamento.delete()
        messages.success(request, 'Planilha importada excluida com sucesso.')
        return redirect('medicoes_obra', obra_id=obra_id)
    return render(
        request,
        'medicoes/confirmar_exclusao_orcamento.html',
        {
            'orcamento': orcamento,
        },
    )


def nova_medicao_construtora(request, orcamento_id):
    orcamento = get_object_or_404(OrcamentoMedicao, id=orcamento_id, obra__empresa=request.empresa)
    initial = {
        'numero': _next_numero(MedicaoConstrutora, orcamento=orcamento),
        'data_medicao': timezone.localdate(),
        'periodo_inicio': timezone.localdate(),
        'periodo_fim': timezone.localdate(),
    }
    if request.method == 'POST':
        form = MedicaoConstrutoraCabecalhoForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                medicao = form.save(commit=False)
                medicao.orcamento = orcamento
                medicao.save()
                ItemMedicaoConstrutora.objects.bulk_create(
                    [
                        ItemMedicaoConstrutora(medicao=medicao, item_orcamento=item)
                        for item in orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_ITEM)
                    ]
                )
            messages.success(request, 'Medicao criada. Agora preencha as quantidades medidas.')
            return redirect('editar_medicao_construtora', medicao_id=medicao.id)
    else:
        form = MedicaoConstrutoraCabecalhoForm(initial=initial)
    return render(request, 'medicoes/form_medicao.html', {'form': form, 'titulo': 'Nova medicao da construtora'})


def editar_medicao_construtora(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoConstrutora.objects.select_related('orcamento', 'orcamento__obra'),
        id=medicao_id,
        orcamento__obra__empresa=request.empresa,
    )
    _sincronizar_itens_medicao_construtora(medicao)
    if request.method == 'POST':
        form = MedicaoConstrutoraForm(request.POST, instance=medicao)
        formset = ItemMedicaoConstrutoraFormSet(request.POST, instance=medicao)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            _sync_faturamentos_diretos(medicao, request.POST)
            _aplicar_percentuais_construtora(medicao)
            messages.success(request, 'Medicao da construtora atualizada com sucesso.')
            return redirect('editar_medicao_construtora', medicao_id=medicao.id)
    else:
        form = MedicaoConstrutoraForm(instance=medicao)
        formset = ItemMedicaoConstrutoraFormSet(instance=medicao)
    faturamentos_ja_descontados = FaturamentoDireto.objects.filter(
        obra=medicao.orcamento.obra,
        vinculos_medicao__isnull=False,
    ).exclude(
        vinculos_medicao__medicao=medicao,
    ).distinct().order_by('data_lancamento', 'id')
    return render(
        request,
        'medicoes/editar_medicao_construtora.html',
        {
            'medicao': medicao,
            'form': form,
            'formset': formset,
            'linhas_medicao': _linhas_medicao_construtora_formset(medicao, formset),
            'faturamentos_diretos_linhas': _faturamentos_diretos_context(medicao),
            'faturamentos_ja_descontados': faturamentos_ja_descontados,
        },
    )


def excluir_medicao_construtora(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoConstrutora.objects.select_related('orcamento'),
        id=medicao_id,
        orcamento__obra__empresa=request.empresa,
    )
    orcamento_id = medicao.orcamento_id
    if request.method == 'POST':
        medicao.delete()
        messages.success(request, 'Medicao da construtora excluida com sucesso.')
        return redirect('detalhe_orcamento_medicao', orcamento_id=orcamento_id)
    return render(
        request,
        'medicoes/confirmar_exclusao_medicao.html',
        {
            'titulo': 'Excluir medicao da construtora',
            'descricao': f'Medicao {medicao.numero} - {medicao.orcamento}',
            'voltar_url': 'editar_medicao_construtora',
            'voltar_arg': medicao.id,
        },
    )


def nova_medicao_empreiteiro_simples(request):
    if request.method == 'POST':
        form = MedicaoEmpreiteiroCabecalhoForm(request.POST, empresa=request.empresa)
        formset = ItemMedicaoEmpreiteiroFormSet(request.POST)
        if form.is_valid():
            medicao = form.save(commit=False)
            medicao.tipo = MedicaoEmpreiteiro.TIPO_SIMPLES
            medicao.empresa = request.empresa
            medicao.save()
            _sync_empreiteiro_medicao(medicao)
            formset = ItemMedicaoEmpreiteiroFormSet(request.POST, instance=medicao)
            if formset.is_valid():
                formset.save()
                messages.success(request, 'Medicao simples de empreiteiro criada.')
                return redirect('editar_medicao_empreiteiro', medicao_id=medicao.id)
            medicao.delete()
    else:
        initial = {
            'numero': _next_numero(MedicaoEmpreiteiro, empresa=request.empresa, tipo=MedicaoEmpreiteiro.TIPO_SIMPLES),
            'data_medicao': timezone.localdate(),
            'periodo_inicio': timezone.localdate(),
            'periodo_fim': timezone.localdate(),
        }
        if request.GET.get('obra'):
            initial['obra'] = request.GET['obra']
        form = MedicaoEmpreiteiroCabecalhoForm(initial=initial, empresa=request.empresa)
        formset = ItemMedicaoEmpreiteiroFormSet()
    return render(
        request,
        'medicoes/editar_medicao_empreiteiro_simples.html',
        {
            'form': form,
            'formset': formset,
            'titulo': 'Nova medicao simples de empreiteiro',
            'empreiteiros_json': _empreiteiros_json(request.empresa),
        },
    )


def nova_medicao_empreiteiro_cumulativa(request, orcamento_id):
    orcamento = get_object_or_404(OrcamentoMedicao, id=orcamento_id, obra__empresa=request.empresa)
    initial = {
        'obra': orcamento.obra,
        'numero': _next_numero(MedicaoEmpreiteiro, empresa=request.empresa, orcamento=orcamento),
        'data_medicao': timezone.localdate(),
        'periodo_inicio': timezone.localdate(),
        'periodo_fim': timezone.localdate(),
    }
    ultima_medicao = _medicoes_empreiteiro_empresa(request.empresa).filter(
        orcamento=orcamento,
        tipo=MedicaoEmpreiteiro.TIPO_CUMULATIVA,
    ).select_related('empreiteiro_cadastro').order_by('-numero', '-id').first()
    if ultima_medicao:
        initial.update(
            {
                'obra': ultima_medicao.obra_id or orcamento.obra_id,
                'empreiteiro_cadastro': ultima_medicao.empreiteiro_cadastro_id,
                'empreiteiro': ultima_medicao.empreiteiro,
                'cpf_cnpj': ultima_medicao.cpf_cnpj,
                'pix': ultima_medicao.pix,
            }
        )
    if request.method == 'POST':
        form = MedicaoEmpreiteiroCabecalhoForm(request.POST, empresa=request.empresa)
        if form.is_valid():
            with transaction.atomic():
                medicao = form.save(commit=False)
                medicao.tipo = MedicaoEmpreiteiro.TIPO_CUMULATIVA
                medicao.orcamento = orcamento
                medicao.obra = medicao.obra or orcamento.obra
                medicao.empresa = request.empresa
                medicao.save()
                _sync_empreiteiro_medicao(medicao)
                ItemMedicaoEmpreiteiro.objects.bulk_create(
                    [
                        ItemMedicaoEmpreiteiro(
                            medicao=medicao,
                            item_orcamento=item,
                            item=item.item,
                            descricao=item.descricao,
                            unidade=item.unidade,
                            valor_unitario=item.preco_unitario_total,
                        )
                        for item in orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_ITEM)
                    ]
                )
            messages.success(request, 'Medicao cumulativa criada. Agora preencha as quantidades medidas.')
            return redirect('editar_medicao_empreiteiro', medicao_id=medicao.id)
    else:
        form = MedicaoEmpreiteiroCabecalhoForm(initial=initial, empresa=request.empresa)
    return render(
        request,
        'medicoes/form_medicao.html',
        {
            'form': form,
            'titulo': 'Nova medicao cumulativa de empreiteiro',
            'empreiteiros_json': _empreiteiros_json(request.empresa),
        },
    )


def editar_medicao_empreiteiro(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoEmpreiteiro.objects.select_related('obra', 'orcamento'),
        id=medicao_id,
        empresa=request.empresa,
    )
    if request.method == 'POST':
        form = MedicaoEmpreiteiroForm(request.POST, instance=medicao, empresa=request.empresa)
        formset = ItemMedicaoEmpreiteiroFormSet(request.POST, instance=medicao, orcamento=medicao.orcamento)
        if form.is_valid() and formset.is_valid():
            form.save()
            _sync_empreiteiro_medicao(medicao)
            formset.save()
            _aplicar_percentuais_empreiteiro(medicao)
            messages.success(request, 'Medicao de empreiteiro atualizada com sucesso.')
            return redirect('editar_medicao_empreiteiro', medicao_id=medicao.id)
    else:
        form = MedicaoEmpreiteiroForm(instance=medicao, empresa=request.empresa)
        formset = ItemMedicaoEmpreiteiroFormSet(instance=medicao, orcamento=medicao.orcamento)
    template = (
        'medicoes/editar_medicao_empreiteiro_simples.html'
        if medicao.tipo == MedicaoEmpreiteiro.TIPO_SIMPLES
        else 'medicoes/editar_medicao_empreiteiro.html'
    )
    return render(
        request,
        template,
        {
            'medicao': medicao,
            'form': form,
            'formset': formset,
            'titulo': 'Medicao de empreiteiro',
            'empreiteiros_json': _empreiteiros_json(request.empresa),
        },
    )


def excluir_medicao_empreiteiro(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoEmpreiteiro.objects.select_related('orcamento'),
        id=medicao_id,
        empresa=request.empresa,
    )
    orcamento_id = medicao.orcamento_id
    if request.method == 'POST':
        medicao.delete()
        messages.success(request, 'Medicao de empreiteiro excluida com sucesso.')
        if orcamento_id:
            return redirect('detalhe_orcamento_medicao', orcamento_id=orcamento_id)
        return redirect('medicoes_empreiteiros_home')
    return render(
        request,
        'medicoes/confirmar_exclusao_medicao.html',
        {
            'titulo': 'Excluir medicao de empreiteiro',
            'descricao': f'Medicao {medicao.numero} - {medicao.empreiteiro}',
            'voltar_url': 'editar_medicao_empreiteiro',
            'voltar_arg': medicao.id,
        },
    )


def _linhas_pdf_medicao(medicao, itens, titulo):
    lines = [
        titulo.upper(),
        f'Emitido em {date.today().strftime("%d/%m/%Y")}',
        f'Obra: {getattr(getattr(medicao, "orcamento", None), "obra", None) or medicao.obra or "-"}',
        f'Medicao: {medicao.numero} | Periodo: {medicao.periodo_inicio:%d/%m/%Y} a {medicao.periodo_fim:%d/%m/%Y}',
        '',
        'ITEM | DESCRICAO | UND | CONTRATO | ANT. | PERIODO | ATUAL | SALDO | VALOR',
    ]
    for item in itens:
        contrato = getattr(getattr(item, 'item_orcamento', None), 'quantidade', Decimal('0'))
        lines.append(
            ' | '.join(
                [
                    (getattr(getattr(item, 'item_orcamento', None), 'item', '') or item.item)[:8],
                    item.descricao[:24] if hasattr(item, 'descricao') else item.item_orcamento.descricao[:24],
                    (item.unidade if hasattr(item, 'unidade') else item.item_orcamento.unidade)[:5],
                    f'{contrato:.4f}',
                    f'{item.quantidade_acumulada_anterior:.4f}',
                    f'{item.quantidade_periodo:.4f}',
                    f'{item.quantidade_acumulada_atual:.4f}',
                    f'{item.saldo_quantidade:.4f}',
                    _money(item.valor_periodo),
                ]
            )
        )
    lines.extend(
        [
            '',
            f'Subtotal do periodo: {_money(medicao.subtotal_periodo)}',
            f'Retencao tecnica: {_money(medicao.retencao_tecnica_calculada if isinstance(medicao, MedicaoConstrutora) else medicao.retencao_tecnica)}',
        ]
    )
    if isinstance(medicao, MedicaoConstrutora):
        lines.extend(
            [
                f'ISSQN: {_money(medicao.issqn_calculado)}',
                f'INSS: {_money(medicao.inss_calculado)}',
                f'Faturamento direto descontado: {_money(medicao.total_faturamento_direto)}',
                f'Base de impostos: {_money(medicao.base_impostos)}',
                f'Base INSS: {_money(medicao.base_inss)}',
            ]
        )
    desconto = medicao.desconto_adicional_calculado if isinstance(medicao, MedicaoConstrutora) else medicao.desconto_adicional
    lines.extend([f'Desconto adicional: {_money(desconto)}', f'Total liquido: {_money(medicao.total_liquido)}'])
    pages = [lines[i : i + 30] for i in range(0, len(lines), 30)] or [[]]
    return _build_simple_pdf(pages)


def medicao_construtora_pdf(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoConstrutora.objects.select_related('orcamento', 'orcamento__obra'),
        id=medicao_id,
        orcamento__obra__empresa=request.empresa,
    )
    response = HttpResponse(
        _pdf_medicao_construtora(medicao),
        content_type='application/pdf',
    )
    response['Content-Disposition'] = f'inline; filename="medicao_construtora_{medicao.numero}.pdf"'
    return response


def medicao_empreiteiro_pdf(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoEmpreiteiro.objects.select_related('obra', 'orcamento'),
        id=medicao_id,
        empresa=request.empresa,
    )
    response = HttpResponse(_pdf_medicao_empreiteiro(medicao), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="medicao_empreiteiro_{medicao.numero}.pdf"'
    return response


def _pdf_medicao_empreiteiro(medicao):
    itens = list(medicao.itens.select_related('item_orcamento'))
    doc = PdfDocument(
        empresa=medicao.empresa,
        title='Boletim de medicao de empreiteiro',
        subtitle=f'Medicao no {medicao.numero}',
        orientation='portrait',
        filename=f'medicao_empreiteiro_{medicao.numero}.pdf',
    )
    doc.add_title(emitted_on=date.today())
    doc.add_info_grid(
        [
            ('Empreiteiro', medicao.empreiteiro),
            ('CPF/CNPJ', medicao.cpf_cnpj or '-'),
            ('PIX', medicao.pix or '-'),
            ('Obra', medicao.obra or getattr(medicao.orcamento, 'obra', '-') or '-'),
            ('Periodo', f'{format_date_br(medicao.periodo_inicio)} a {format_date_br(medicao.periodo_fim)}'),
            ('Data da medicao', format_date_br(medicao.data_medicao)),
        ],
        columns=3,
    )
    if medicao.tipo == MedicaoEmpreiteiro.TIPO_SIMPLES:
        columns = [
            PdfTableColumn('item', 'Item', weight=0.8, align='center'),
            PdfTableColumn('descricao', 'Descricao', weight=3.9),
            PdfTableColumn('unidade', 'Und', weight=0.75, align='center'),
            PdfTableColumn('quantidade', 'Qtd', weight=0.9, align='right'),
            PdfTableColumn('valor_unitario', 'Valor unit.', weight=1.2, align='right'),
            PdfTableColumn('total', 'Total', weight=1.25, align='right'),
        ]
        rows = [
            {
                'item': item.item or '-',
                'descricao': item.descricao,
                'unidade': item.unidade or '-',
                'quantidade': format_decimal_br(item.quantidade_periodo, 4),
                'valor_unitario': _money(item.valor_unitario),
                'total': _money(item.valor_periodo),
            }
            for item in itens
        ]
    else:
        columns = [
            PdfTableColumn('item', 'Item', weight=0.7, align='center'),
            PdfTableColumn('descricao', 'Descricao', weight=3.5),
            PdfTableColumn('unidade', 'Und', weight=0.65, align='center'),
            PdfTableColumn('anterior', 'Anterior', weight=0.95, align='right'),
            PdfTableColumn('periodo', 'Periodo', weight=0.95, align='right'),
            PdfTableColumn('atual', 'Atual', weight=0.95, align='right'),
            PdfTableColumn('saldo', 'Saldo', weight=0.95, align='right'),
            PdfTableColumn('total', 'Total', weight=1.25, align='right'),
        ]
        rows = [
            {
                'item': item.item or '-',
                'descricao': item.descricao,
                'unidade': item.unidade or '-',
                'anterior': format_decimal_br(item.quantidade_acumulada_anterior, 4),
                'periodo': format_decimal_br(item.quantidade_periodo, 4),
                'atual': format_decimal_br(item.quantidade_acumulada_atual, 4),
                'saldo': format_decimal_br(item.saldo_quantidade, 4),
                'total': _money(item.valor_periodo),
            }
            for item in itens
        ]
    doc.add_table(columns, rows, row_height=44)
    doc.add_totals_box(
        [
            ('Subtotal medido', _money(medicao.subtotal_periodo), False),
            ('Retencao tecnica', f'- {_money(medicao.retencao_tecnica)}', False),
            ('Desconto adicional', f'- {_money(medicao.desconto_adicional)}', False),
            ('Total liquido', _money(medicao.total_liquido), True),
        ],
        width=620,
    )
    if medicao.observacoes:
        doc.add_section_header('Observacoes')
        doc.add_info_grid([('Observacoes', medicao.observacoes)], columns=1)
    return doc.build()


def _xlsx_medicao(medicao, itens):
    empresa = medicao.orcamento.obra.empresa if isinstance(medicao, MedicaoConstrutora) else medicao.empresa
    builder = ExcelReportBuilder(
        empresa=empresa,
        title='Boletim de medicao',
        subtitle=f'Medicao no {medicao.numero} | Periodo {format_date_br(medicao.periodo_inicio)} a {format_date_br(medicao.periodo_fim)}',
        sheet_name='Medicao',
        orientation='landscape',
    )
    builder.add_header(emitted_on=date.today())

    if isinstance(medicao, MedicaoConstrutora):
        columns = [
            ExcelColumn('item', 'Item', width=8, align='center'),
            ExcelColumn('descricao', 'Descricao', width=56, wrap=True),
            ExcelColumn('unidade', 'Unidade', width=9, align='center'),
            ExcelColumn('quantidade', 'Qtde', width=11, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('unit_material', 'Unit. material', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('unit_mao_obra', 'Unit. mao obra', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('unit_equip', 'Unit. equip.', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('preco_unit', 'Unit. total', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('anterior', 'Acum. anterior', width=14, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('periodo', 'Periodo', width=12, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('percentual', '% executado', width=12, align='center', number_format=ExcelReportBuilder.PERCENT_FORMAT),
            ExcelColumn('material', 'Material', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('mao_obra', 'Mao de obra', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('equip', 'Equip.', width=16, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
            ExcelColumn('valor', 'Valor total', width=17, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ]
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        group_row = builder.current_row
        for start, end, label, fill in [
            (1, 8, 'Itens contratuais', 'contract'),
            (9, 11, 'Itens medidos', 'measured'),
            (12, 15, 'Valor a receber', 'receivable'),
        ]:
            builder.ws.merge_cells(start_row=group_row, start_column=start, end_row=group_row, end_column=end)
            cell = builder.ws.cell(group_row, start, label)
            cell.fill = builder._fill(fill)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col in range(start, end + 1):
                current = builder.ws.cell(group_row, col)
                current.fill = builder._fill(fill)
                current.border = border
        builder.current_row += 1
    else:
        columns = [
            ExcelColumn('item', 'Item', width=10, align='center'),
            ExcelColumn('descricao', 'Descricao', width=58, wrap=True),
            ExcelColumn('unidade', 'Unidade', width=12, align='center'),
            ExcelColumn('contrato', 'Contrato', width=14, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('anterior', 'Acum. anterior', width=16, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('periodo', 'Periodo', width=14, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('atual', 'Acum. atual', width=14, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('saldo', 'Saldo', width=14, align='center', number_format=ExcelReportBuilder.DECIMAL_FORMAT),
            ExcelColumn('valor', 'Valor', width=17, align='right', number_format=ExcelReportBuilder.MONEY_FORMAT),
        ]
    rows = []
    for item in itens:
        if isinstance(item, ItemOrcamentoMedicao) and item.eh_grupo:
            rows.append({'__merge_label__': f'{item.item} - {item.descricao}', '__bg': 'surface'})
            continue
        contrato = getattr(getattr(item, 'item_orcamento', None), 'quantidade', Decimal('0'))
        base = getattr(item, 'item_orcamento', None)
        if isinstance(medicao, MedicaoConstrutora):
            rows.append(
                {
                    'item': base.item,
                    'descricao': base.descricao,
                    'unidade': base.unidade or '-',
                    'quantidade': base.quantidade,
                    'unit_material': base.preco_unitario_material,
                    'unit_mao_obra': base.preco_unitario_mao_obra,
                    'unit_equip': base.preco_unitario_equipamentos,
                    'preco_unit': base.preco_unitario_total,
                    'anterior': item.quantidade_acumulada_anterior,
                    'periodo': item.quantidade_periodo,
                    'percentual': _percent_from_item(item) / Decimal('100'),
                    'material': item.valor_material_periodo,
                    'mao_obra': item.valor_mao_obra_periodo,
                    'equip': item.valor_equipamentos_periodo,
                    'valor': item.valor_periodo,
                    '__cell_bgs': {'anterior': 'measured', 'periodo': 'measured', 'percentual': 'measured'},
                }
            )
        else:
            rows.append(
                {
                    'item': getattr(base, 'item', '') or getattr(item, 'item', ''),
                    'descricao': item.descricao if hasattr(item, 'descricao') else base.descricao,
                    'unidade': item.unidade if hasattr(item, 'unidade') else base.unidade,
                    'contrato': contrato,
                    'anterior': item.quantidade_acumulada_anterior,
                    'periodo': item.quantidade_periodo,
                    'atual': item.quantidade_acumulada_atual,
                    'saldo': item.saldo_quantidade,
                    'valor': item.valor_periodo,
                }
            )
    if isinstance(medicao, MedicaoConstrutora):
        rows.extend(
            [
                {'descricao': 'Total material', 'material': medicao.total_material_periodo, 'valor': medicao.total_material_periodo, '__bg': 'surface', '__bold': True},
                {'descricao': 'Total mao de obra', 'mao_obra': medicao.total_mao_obra_periodo, 'valor': medicao.total_mao_obra_periodo, '__bg': 'surface', '__bold': True},
                {'descricao': 'Total equipamentos', 'equip': medicao.total_equipamentos_periodo, 'valor': medicao.total_equipamentos_periodo, '__bg': 'surface', '__bold': True},
            ]
        )
    table_header_row = builder.current_row
    builder.add_table(columns, rows)
    if isinstance(medicao, MedicaoConstrutora):
        for col in range(1, 16):
            if col <= 8:
                fill = 'contract'
            elif col <= 11:
                fill = 'measured'
            else:
                fill = 'receivable'
            builder.ws.cell(table_header_row, col).fill = builder._fill(fill)

    ws = builder.ws
    summary_start = ws.max_row + 1
    ws.append(['Resumo'])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
    ws.append(['Valor bruto', medicao.total_bruto if isinstance(medicao, MedicaoConstrutora) else medicao.subtotal_periodo])
    if isinstance(medicao, MedicaoConstrutora):
        ws.append(['Total material medido', medicao.total_material_periodo])
        ws.append(['Total mao de obra medida', medicao.total_mao_obra_periodo])
        ws.append(['Total equipamentos medido', medicao.total_equipamentos_periodo])
    ws.append(['Retencao tecnica', medicao.retencao_tecnica_calculada if isinstance(medicao, MedicaoConstrutora) else medicao.retencao_tecnica])
    if isinstance(medicao, MedicaoConstrutora):
        ws.append(['ISSQN', medicao.issqn_calculado])
        ws.append(['INSS', medicao.inss_calculado])
        ws.append(['Faturamento direto descontado', medicao.total_faturamento_direto])
        ws.append(['Base de impostos', medicao.base_impostos])
        ws.append(['Material para NF', medicao.valor_material_nf])
        ws.append(['Mao de obra para NF', medicao.valor_mao_obra_nf])
        ws.append(['Equipamentos para NF', medicao.valor_equipamentos_nf])
        ws.append(['Base INSS', medicao.base_inss])
    ws.append(['Desconto adicional', medicao.desconto_adicional_calculado if isinstance(medicao, MedicaoConstrutora) else medicao.desconto_adicional])
    ws.append(['Total liquido', medicao.total_liquido])
    for row in ws.iter_rows(min_row=summary_start + 1, max_row=ws.max_row, min_col=1, max_col=2):
        row[0].font = Font(bold=row[0].value == 'Total liquido')
        row[1].number_format = ExcelReportBuilder.MONEY_FORMAT
        row[1].font = Font(bold=row[0].value == 'Total liquido')
    return builder.build()


def medicao_construtora_excel(request, medicao_id):
    medicao = get_object_or_404(
        MedicaoConstrutora.objects.select_related('orcamento', 'orcamento__obra'),
        id=medicao_id,
        orcamento__obra__empresa=request.empresa,
    )
    response = HttpResponse(
        _xlsx_medicao(medicao, _itens_medicao_construtora_com_grupos(medicao)),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="medicao_construtora_{medicao.numero}.xlsx"'
    return response


def medicao_empreiteiro_excel(request, medicao_id):
    medicao = get_object_or_404(MedicaoEmpreiteiro, id=medicao_id, empresa=request.empresa)
    response = HttpResponse(
        _xlsx_medicao(medicao, medicao.itens.select_related('item_orcamento')),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="medicao_empreiteiro_{medicao.numero}.xlsx"'
    return response
