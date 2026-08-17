from datetime import date
from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from empresas.models import Empresa, UsuarioEmpresa
from obras.models import Obra
from controles.models import FaturamentoDireto

from .forms import RelatorioMedicoesForm
from .models import (
    Empreiteiro,
    FaturamentoDiretoMedicao,
    ItemMedicaoConstrutora,
    ItemMedicaoEmpreiteiro,
    ItemOrcamentoMedicao,
    MedicaoConstrutora,
    MedicaoEmpreiteiro,
    OrcamentoMedicao,
)


class MedicoesTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(username='usuario', password='senha')
        self.empresa = Empresa.objects.get(slug='ambar')
        UsuarioEmpresa.objects.create(usuario=self.user, empresa=self.empresa)
        self.client.force_login(self.user)
        self.obra = Obra.objects.create(empresa=self.empresa, nome_obra='Obra Teste', cliente='Cliente')

    def _orcamento(self):
        orcamento = OrcamentoMedicao.objects.create(
            obra=self.obra,
            nome='Orcamento principal',
            tipo=OrcamentoMedicao.TIPO_CONSTRUTORA,
        )
        item = ItemOrcamentoMedicao.objects.create(
            orcamento=orcamento,
            item='1.1',
            descricao='Escavacao',
            unidade='m3',
            quantidade=Decimal('100.0000'),
            preco_unitario_material=Decimal('10.00'),
            preco_unitario_mao_obra=Decimal('5.00'),
            preco_unitario_equipamentos=Decimal('2.00'),
        )
        return orcamento, item

    def test_importa_orcamento_csv(self):
        arquivo = SimpleUploadedFile(
            'orcamento.csv',
            (
                'item;descricao;unidade;quantidade;preco unitario material;preco unitario mao de obra;preco unitario equipamentos\n'
                '1;Drenagem;m;10,5;100,1234;20,0001;5,0000\n'
            ).encode('utf-8-sig'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('importar_orcamento_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha da obra',
                'tipo': OrcamentoMedicao.TIPO_CONSTRUTORA,
                'arquivo': arquivo,
            },
        )

        self.assertEqual(response.status_code, 302)
        orcamento = OrcamentoMedicao.objects.get(nome='Planilha da obra')
        item = orcamento.itens.get()
        self.assertEqual(item.quantidade, Decimal('10.5'))
        self.assertEqual(item.preco_unitario_total, Decimal('125.1235'))

    def test_importa_planilha_com_preco_unitario_simples(self):
        arquivo = SimpleUploadedFile(
            'planilha.csv',
            (
                'referencia;descricao;un;quantidade;preco unitario\n'
                '1.1;Servico medido;m2;10;45,50\n'
            ).encode('utf-8-sig'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('importar_orcamento_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha com unitario',
                'tipo': OrcamentoMedicao.TIPO_CONSTRUTORA,
                'arquivo': arquivo,
            },
        )

        self.assertEqual(response.status_code, 302)
        item = OrcamentoMedicao.objects.get(nome='Planilha com unitario').itens.get()
        self.assertEqual(item.preco_unitario_total, Decimal('45.50'))

    def test_importa_planilha_com_grupos_e_itens_mediveis(self):
        arquivo = SimpleUploadedFile(
            'planilha_grupos.csv',
            (
                'item;tipo;descricao;unidade;quantidade;preco_unitario_material;preco_unitario_mao_obra;preco_unitario_equipamentos\n'
                '1;grupo;SERVICOS PRELIMINARES;;;;;\n'
                '1.1;item;Mobilizacao do canteiro;mes;6;0;3500;0\n'
                '2;;TERRAPLANAGEM;;0;0;0;0\n'
                '2.1;;Escavacao;m3;10;1;2;3\n'
            ).encode('utf-8-sig'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('importar_orcamento_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha com grupos',
                'tipo': OrcamentoMedicao.TIPO_CONSTRUTORA,
                'arquivo': arquivo,
            },
        )

        self.assertEqual(response.status_code, 302)
        orcamento = OrcamentoMedicao.objects.get(nome='Planilha com grupos')
        self.assertEqual(orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_GRUPO).count(), 2)
        self.assertEqual(orcamento.itens.filter(tipo=ItemOrcamentoMedicao.TIPO_ITEM).count(), 2)
        self.assertEqual(orcamento.total_orcamento, Decimal('21060.0000'))

    def test_tela_medicao_mostra_detalhes_e_historico_do_faturamento_direto(self):
        orcamento, item = self._orcamento()
        medicao_atual = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=2,
            periodo_inicio=date(2026, 2, 1),
            periodo_fim=date(2026, 2, 28),
            data_medicao=date(2026, 2, 28),
        )
        medicao_anterior = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        disponivel = FaturamentoDireto.objects.create(
            obra=self.obra,
            numero_nf='1414',
            empresa_comprou='Fornecedor livre',
            valor_nota=Decimal('1000.00'),
            descricao='Tubos de concreto',
            vencimento_boleto='27/04/2026',
        )
        usado = FaturamentoDireto.objects.create(
            obra=self.obra,
            numero_nf='1413',
            empresa_comprou='Fornecedor usado',
            valor_nota=Decimal('500.00'),
            descricao='Barras de aco',
            vencimento_boleto='24/04/2026',
            medicao_desconto='Medicao 1',
        )
        FaturamentoDiretoMedicao.objects.create(medicao=medicao_anterior, faturamento_direto=usado)

        response = self.client.get(reverse('editar_medicao_construtora', args=[medicao_atual.id]))

        self.assertContains(response, '1414')
        self.assertContains(response, 'Fornecedor livre')
        self.assertContains(response, 'R$ 1.000,00')
        self.assertContains(response, '100,00%')
        self.assertContains(response, 'Historico de faturamento direto ja descontado')
        self.assertContains(response, 'Fornecedor usado')

    def test_importacao_sem_cabecalho_retorna_erro_no_formulario(self):
        arquivo = SimpleUploadedFile(
            'orcamento.csv',
            '1;Drenagem;m;10,5\n'.encode('utf-8'),
            content_type='text/csv',
        )

        response = self.client.post(
            reverse('importar_orcamento_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha invalida',
                'tipo': OrcamentoMedicao.TIPO_CONSTRUTORA,
                'arquivo': arquivo,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(OrcamentoMedicao.objects.filter(nome='Planilha invalida').exists())

    def test_edita_itens_da_planilha_importada(self):
        orcamento, item = self._orcamento()

        response = self.client.post(
            reverse('editar_itens_orcamento_medicao', args=[orcamento.id]),
            {
                'itens-TOTAL_FORMS': '2',
                'itens-INITIAL_FORMS': '1',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': str(item.id),
                'itens-0-tipo': ItemOrcamentoMedicao.TIPO_ITEM,
                'itens-0-ordem': '0',
                'itens-0-item': '1.1',
                'itens-0-descricao': 'Escavacao revisada',
                'itens-0-unidade': 'm3',
                'itens-0-quantidade': '120.1234',
                'itens-0-preco_unitario_material': '10.1111',
                'itens-0-preco_unitario_mao_obra': '5.2222',
                'itens-0-preco_unitario_equipamentos': '2.3333',
                'itens-1-id': '',
                'itens-1-tipo': ItemOrcamentoMedicao.TIPO_ITEM,
                'itens-1-ordem': '1',
                'itens-1-item': '1.2',
                'itens-1-descricao': 'Transporte',
                'itens-1-unidade': 'm3',
                'itens-1-quantidade': '10.0000',
                'itens-1-preco_unitario_material': '1.0000',
                'itens-1-preco_unitario_mao_obra': '0.0000',
                'itens-1-preco_unitario_equipamentos': '0.0000',
            },
        )

        self.assertRedirects(response, reverse('detalhe_orcamento_medicao', args=[orcamento.id]))
        item.refresh_from_db()
        self.assertEqual(item.descricao, 'Escavacao revisada')
        self.assertEqual(item.quantidade, Decimal('120.1234'))
        self.assertEqual(item.preco_unitario_total, Decimal('17.6666'))
        self.assertEqual(orcamento.itens.count(), 2)

    def test_insere_grupo_antes_do_primeiro_item_da_planilha(self):
        orcamento, item = self._orcamento()

        response = self.client.post(
            reverse('editar_itens_orcamento_medicao', args=[orcamento.id]),
            {
                'itens-TOTAL_FORMS': '2',
                'itens-INITIAL_FORMS': '1',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': str(item.id),
                'itens-0-tipo': ItemOrcamentoMedicao.TIPO_ITEM,
                'itens-0-ordem': '1',
                'itens-0-item': '1.1',
                'itens-0-descricao': 'Escavacao',
                'itens-0-unidade': 'm3',
                'itens-0-quantidade': '100.0000',
                'itens-0-preco_unitario_material': '10.0000',
                'itens-0-preco_unitario_mao_obra': '5.0000',
                'itens-0-preco_unitario_equipamentos': '2.0000',
                'itens-1-id': '',
                'itens-1-tipo': ItemOrcamentoMedicao.TIPO_GRUPO,
                'itens-1-ordem': '0',
                'itens-1-item': '1',
                'itens-1-descricao': 'TERRAPLANAGEM',
                'itens-1-unidade': '',
                'itens-1-quantidade': '',
                'itens-1-preco_unitario_material': '',
                'itens-1-preco_unitario_mao_obra': '',
                'itens-1-preco_unitario_equipamentos': '',
            },
        )

        self.assertRedirects(response, reverse('detalhe_orcamento_medicao', args=[orcamento.id]))
        itens = list(orcamento.itens.order_by('ordem', 'id'))
        self.assertEqual(itens[0].tipo, ItemOrcamentoMedicao.TIPO_GRUPO)
        self.assertEqual(itens[0].descricao, 'TERRAPLANAGEM')
        self.assertEqual(itens[1].id, item.id)
        self.assertEqual(orcamento.total_orcamento, Decimal('1700.00000000'))

    def test_saldo_contratual_construtora_mostra_somente_itens_com_saldo(self):
        orcamento, item = self._orcamento()
        item_medido_total = ItemOrcamentoMedicao.objects.create(
            orcamento=orcamento,
            item='2.1',
            descricao='Servico ja concluido',
            unidade='m',
            quantidade=Decimal('10.0000'),
            preco_unitario_material=Decimal('1.0000'),
            preco_unitario_mao_obra=Decimal('1.0000'),
            preco_unitario_equipamentos=Decimal('0.0000'),
        )
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        ItemMedicaoConstrutora.objects.create(
            medicao=medicao,
            item_orcamento=item,
            quantidade_periodo=Decimal('40.0000'),
        )
        ItemMedicaoConstrutora.objects.create(
            medicao=medicao,
            item_orcamento=item_medido_total,
            quantidade_periodo=Decimal('10.0000'),
        )

        response = self.client.get(reverse('saldo_contratual_construtora', args=[orcamento.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Escavacao')
        self.assertContains(response, '60,0000')
        self.assertContains(response, 'R$ 1.020,00')
        self.assertNotContains(response, 'Servico ja concluido')

        pdf = self.client.get(reverse('saldo_contratual_construtora_pdf', args=[orcamento.id]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))

        excel = self.client.get(reverse('saldo_contratual_construtora_excel', args=[orcamento.id]))
        self.assertEqual(excel.status_code, 200)
        self.assertIn('spreadsheetml', excel['Content-Type'])
        wb = load_workbook(BytesIO(excel.content))
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'Saldo contratual da construtora')
        self.assertIn('Escavacao', [cell.value for cell in ws['B']])

    def test_cria_planilha_manual_e_medicao_construtora(self):
        response = self.client.post(
            reverse('novo_orcamento_manual_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha manual construtora',
                'tipo': OrcamentoMedicao.TIPO_CONSTRUTORA,
                'observacoes': '',
            },
        )

        orcamento = OrcamentoMedicao.objects.get(nome='Planilha manual construtora')
        self.assertRedirects(response, reverse('editar_itens_orcamento_medicao', args=[orcamento.id]))

        self.client.post(
            reverse('editar_itens_orcamento_medicao', args=[orcamento.id]),
            {
                'itens-TOTAL_FORMS': '1',
                'itens-INITIAL_FORMS': '0',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': '',
                'itens-0-tipo': ItemOrcamentoMedicao.TIPO_ITEM,
                'itens-0-ordem': '0',
                'itens-0-item': '1',
                'itens-0-descricao': 'Servico manual',
                'itens-0-unidade': 'm2',
                'itens-0-quantidade': '10.0000',
                'itens-0-preco_unitario_material': '2.0000',
                'itens-0-preco_unitario_mao_obra': '3.0000',
                'itens-0-preco_unitario_equipamentos': '0.0000',
            },
        )

        response = self.client.post(
            reverse('nova_medicao_construtora', args=[orcamento.id]),
            {
                'numero': '1',
                'periodo_inicio': '2026-01-01',
                'periodo_fim': '2026-01-31',
                'data_medicao': '2026-01-31',
                'observacoes': '',
            },
        )

        medicao = MedicaoConstrutora.objects.get(orcamento=orcamento)
        self.assertRedirects(response, reverse('editar_medicao_construtora', args=[medicao.id]))
        self.assertEqual(medicao.itens.count(), 1)

    def test_cria_planilha_manual_e_medicao_cumulativa_empreiteiro(self):
        empreiteiro = Empreiteiro.objects.create(
            empresa=self.empresa,
            nome='Empreiteiro cadastrado',
            cpf_cnpj='11.111.111/0001-11',
            pix='pix@empreiteiro.com',
        )
        response = self.client.post(
            reverse('novo_orcamento_manual_medicao'),
            {
                'obra': self.obra.id,
                'nome': 'Planilha manual empreiteiro',
                'tipo': OrcamentoMedicao.TIPO_EMPREITEIRO,
                'observacoes': '',
            },
        )

        orcamento = OrcamentoMedicao.objects.get(nome='Planilha manual empreiteiro')
        self.assertRedirects(response, reverse('editar_itens_orcamento_medicao', args=[orcamento.id]))

        self.client.post(
            reverse('editar_itens_orcamento_medicao', args=[orcamento.id]),
            {
                'itens-TOTAL_FORMS': '1',
                'itens-INITIAL_FORMS': '0',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': '',
                'itens-0-tipo': ItemOrcamentoMedicao.TIPO_ITEM,
                'itens-0-ordem': '0',
                'itens-0-item': '1',
                'itens-0-descricao': 'Servico empreiteiro manual',
                'itens-0-unidade': 'm',
                'itens-0-quantidade': '20.0000',
                'itens-0-preco_unitario_material': '0.0000',
                'itens-0-preco_unitario_mao_obra': '15.0000',
                'itens-0-preco_unitario_equipamentos': '0.0000',
            },
        )

        response = self.client.post(
            reverse('nova_medicao_empreiteiro_cumulativa', args=[orcamento.id]),
            {
                'obra': self.obra.id,
                'empreiteiro_cadastro': empreiteiro.id,
                'empreiteiro': '',
                'cpf_cnpj': '',
                'pix': '',
                'numero': '1',
                'periodo_inicio': '2026-02-01',
                'periodo_fim': '2026-02-28',
                'data_medicao': '2026-02-28',
                'observacoes': '',
            },
        )

        medicao = MedicaoEmpreiteiro.objects.get(orcamento=orcamento)
        self.assertRedirects(response, reverse('editar_medicao_empreiteiro', args=[medicao.id]))
        self.assertEqual(medicao.tipo, MedicaoEmpreiteiro.TIPO_CUMULATIVA)
        self.assertEqual(medicao.empreiteiro_cadastro, empreiteiro)
        self.assertEqual(medicao.empreiteiro, 'Empreiteiro cadastrado')
        self.assertEqual(medicao.itens.count(), 1)
        item_medicao = medicao.itens.get()
        item_medicao.quantidade_periodo = Decimal('5.0000')
        item_medicao.save()

        response_obra = self.client.get(reverse('medicoes_obra', args=[self.obra.id]))

        self.assertContains(response_obra, 'Pago/Liquido')
        self.assertContains(response_obra, '% concluida')
        self.assertContains(response_obra, 'R$ 300,00')
        self.assertContains(response_obra, 'R$ 75,00')
        self.assertContains(response_obra, 'R$ 225,00')
        self.assertContains(response_obra, '25,00%')

        response_detalhe = self.client.get(reverse('detalhe_orcamento_medicao', args=[orcamento.id]))
        self.assertContains(response_detalhe, 'Contratado')
        self.assertContains(response_detalhe, 'Medido')
        self.assertContains(response_detalhe, 'Saldo contratual')
        self.assertContains(response_detalhe, 'R$ 300,00')
        self.assertContains(response_detalhe, 'R$ 75,00')
        self.assertContains(response_detalhe, 'R$ 225,00')
        self.assertContains(response_detalhe, '25,00% concluida')

        response_home = self.client.get(reverse('medicoes_empreiteiros_home'))
        self.assertContains(response_home, 'R$ 300,00')
        self.assertContains(response_home, 'R$ 75,00')
        self.assertContains(response_home, 'R$ 225,00')
        self.assertContains(response_home, '25,00%')

    def test_relatorio_gerencial_medicoes_filtra_colunas_e_exporta(self):
        orcamento, item = self._orcamento()
        medicao_construtora = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 3, 1),
            periodo_fim=date(2026, 3, 31),
            data_medicao=date(2026, 3, 31),
        )
        ItemMedicaoConstrutora.objects.create(
            medicao=medicao_construtora,
            item_orcamento=item,
            quantidade_periodo=Decimal('10.0000'),
        )
        empreiteiro = Empreiteiro.objects.create(empresa=self.empresa, nome='Empreiteiro Relatorio')
        medicao_empreiteiro = MedicaoEmpreiteiro.objects.create(
            empresa=self.empresa,
            tipo=MedicaoEmpreiteiro.TIPO_SIMPLES,
            obra=self.obra,
            empreiteiro_cadastro=empreiteiro,
            empreiteiro='Empreiteiro Relatorio',
            numero=2,
            periodo_inicio=date(2026, 4, 1),
            periodo_fim=date(2026, 4, 30),
            data_medicao=date(2026, 4, 30),
        )
        ItemMedicaoEmpreiteiro.objects.create(
            medicao=medicao_empreiteiro,
            item='1',
            descricao='Servico relatorio',
            quantidade_periodo=Decimal('2.0000'),
            valor_unitario=Decimal('50.00'),
        )
        for index in range(16):
            extra = MedicaoEmpreiteiro.objects.create(
                empresa=self.empresa,
                tipo=MedicaoEmpreiteiro.TIPO_SIMPLES,
                obra=self.obra,
                empreiteiro_cadastro=empreiteiro,
                empreiteiro='Empreiteiro Relatorio',
                numero=index + 3,
                periodo_inicio=date(2026, 4, 1),
                periodo_fim=date(2026, 4, 30),
                data_medicao=date(2026, 4, 30),
            )
            ItemMedicaoEmpreiteiro.objects.create(
                medicao=extra,
                item='1',
                descricao='Servico relatorio extra',
                quantidade_periodo=Decimal('1.0000'),
                valor_unitario=Decimal('10.00'),
            )

        response = self.client.get(
            reverse('relatorio_medicoes'),
            {
                'tipo': 'empreiteiro',
                'empreiteiro': empreiteiro.id,
                'colunas': ['tipo', 'empreiteiro', 'data_medicao', 'medido', 'liquido'],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Relatorio gerencial de medicoes')
        self.assertContains(response, 'Empreiteiro Relatorio')
        self.assertContains(response, '30/04/2026')
        self.assertContains(response, 'R$ 100,00')
        self.assertNotContains(response, 'R$ 170,00')

        excel = self.client.get(
            reverse('relatorio_medicoes'),
            {
                'tipo': 'empreiteiro',
                'empreiteiro': empreiteiro.id,
                'colunas': ['tipo', 'empreiteiro', 'data_medicao', 'medido'],
                'export': 'excel',
            },
        )
        self.assertEqual(excel.status_code, 200)
        self.assertIn('spreadsheetml', excel['Content-Type'])
        wb = load_workbook(BytesIO(excel.content))
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'Relatorio gerencial de medicoes')
        header_row = next(row for row in ws.iter_rows(values_only=True) if row and row[0] == 'Tipo')
        self.assertEqual(header_row[:4], ('Tipo', 'Empreiteiro', 'Data da medicao', 'Valor medido'))
        empreiteiro_row = next(row for row in ws.iter_rows(values_only=True) if row and row[1] == 'Empreiteiro Relatorio')
        self.assertEqual(empreiteiro_row[1], 'Empreiteiro Relatorio')
        self.assertTrue(ws.auto_filter.ref)
        self.assertTrue(ws.freeze_panes)

        pdf = self.client.get(
            reverse('relatorio_medicoes'),
            {
                'tipo': 'empreiteiro',
                'empreiteiro': empreiteiro.id,
                'colunas': [choice[0] for choice in RelatorioMedicoesForm.COLUNAS_CHOICES],
                'export': 'pdf',
            },
        )
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertTrue(pdf.content.startswith(b'%PDF'))

    def test_exclui_planilha_importada(self):
        orcamento, item = self._orcamento()

        response = self.client.post(reverse('excluir_orcamento_medicao', args=[orcamento.id]))

        self.assertRedirects(response, reverse('medicoes_obra', args=[self.obra.id]))
        self.assertFalse(OrcamentoMedicao.objects.filter(id=orcamento.id).exists())
        self.assertFalse(ItemOrcamentoMedicao.objects.filter(id=item.id).exists())

    def test_abre_formulario_medicao_simples(self):
        response = self.client.get(reverse('nova_medicao_empreiteiro_simples'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Recibo simples de medição')
        self.assertContains(response, 'Adicionar item')
        self.assertContains(response, 'Buscar por nome, CPF/CNPJ ou PIX')
        self.assertNotContains(response, 'Empreiteiro novo/manual')
        self.assertContains(response, 'name="itens-TOTAL_FORMS" value="0"')

    def test_telas_separadas_de_medicao_carregam(self):
        orcamento, item = self._orcamento()
        response_construtora = self.client.get(reverse('medicoes_construtora_home'))
        response_empreiteiros = self.client.get(reverse('medicoes_empreiteiros_home'))
        response_obra = self.client.get(reverse('medicoes_obra', args=[self.obra.id]))

        self.assertContains(response_construtora, 'Medição da construtora')
        self.assertContains(response_construtora, orcamento.nome)
        self.assertContains(response_empreiteiros, 'Medição de empreiteiro')
        self.assertContains(response_obra, 'Painel operacional da obra')

    def test_medicao_construtora_calcula_acumulado_e_liquido(self):
        orcamento, item = self._orcamento()
        primeira = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        ItemMedicaoConstrutora.objects.create(medicao=primeira, item_orcamento=item, quantidade_periodo=Decimal('20'))
        segunda = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=2,
            periodo_inicio=date(2026, 2, 1),
            periodo_fim=date(2026, 2, 28),
            data_medicao=date(2026, 2, 28),
            retencao_tecnica=Decimal('10.00'),
            issqn=Decimal('5.00'),
            inss=Decimal('3.00'),
            desconto_adicional=Decimal('2.00'),
        )
        item_segunda = ItemMedicaoConstrutora.objects.create(
            medicao=segunda,
            item_orcamento=item,
            quantidade_periodo=Decimal('30'),
        )

        self.assertEqual(item_segunda.quantidade_acumulada_anterior, Decimal('20'))
        self.assertEqual(item_segunda.quantidade_acumulada_atual, Decimal('50'))
        self.assertEqual(segunda.subtotal_periodo, Decimal('510.00'))
        self.assertEqual(segunda.total_liquido, Decimal('490.00'))

        pdf = self.client.get(reverse('medicao_construtora_pdf', args=[segunda.id]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')

        excel = self.client.get(reverse('medicao_construtora_excel', args=[segunda.id]))
        wb = load_workbook(BytesIO(excel.content))
        rows = list(wb.active.iter_rows(values_only=True))
        self.assertTrue(any(row and row[0] == 'Itens contratuais' for row in rows))
        headers = next(row for row in rows if row and row[0] == 'Item')
        self.assertIn('Unit. material', headers)
        self.assertIn('Unit. mao obra', headers)
        self.assertIn('Material', headers)
        self.assertIn('Mao de obra', headers)
        self.assertIn('Equip.', headers)

    def test_medicao_construtora_salva_mesmo_com_grupo_antigo_na_medicao(self):
        orcamento, item = self._orcamento()
        grupo = ItemOrcamentoMedicao.objects.create(
            orcamento=orcamento,
            tipo=ItemOrcamentoMedicao.TIPO_GRUPO,
            item='1',
            descricao='SERVICOS PRELIMINARES',
        )
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=grupo)
        item_medicao = ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item)

        response = self.client.post(
            reverse('editar_medicao_construtora', args=[medicao.id]),
            {
                'numero': '1',
                'periodo_inicio': '2026-01-01',
                'periodo_fim': '2026-01-31',
                'data_medicao': '2026-01-31',
                'retencao_tecnica': '0',
                'retencao_tecnica_percentual': '0',
                'issqn': '0',
                'issqn_percentual': '0',
                'inss': '0',
                'inss_percentual': '0',
                'desconto_adicional': '0',
                'desconto_adicional_percentual': '0',
                'observacoes': '',
                'itens-TOTAL_FORMS': '1',
                'itens-INITIAL_FORMS': '1',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': str(item_medicao.id),
                'itens-0-quantidade_periodo': '15',
            },
        )

        self.assertRedirects(response, reverse('editar_medicao_construtora', args=[medicao.id]))
        item_medicao.refresh_from_db()
        self.assertEqual(item_medicao.quantidade_periodo, Decimal('15'))
        self.assertFalse(medicao.itens.filter(item_orcamento=grupo).exists())

    def test_medicao_construtora_desconta_faturamento_direto_fora_da_base_de_impostos(self):
        orcamento, item = self._orcamento()
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
            issqn_percentual=Decimal('5.00'),
            inss_percentual=Decimal('10.00'),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item, quantidade_periodo=Decimal('20'))
        faturamento = FaturamentoDireto.objects.create(
            obra=self.obra,
            numero_nf='FD-1',
            empresa_comprou='Cliente',
            valor_nota=Decimal('100.00'),
            descricao='Material comprado direto',
            vencimento_boleto='30 dias',
        )
        FaturamentoDiretoMedicao.objects.create(medicao=medicao, faturamento_direto=faturamento)

        self.assertEqual(medicao.total_bruto, Decimal('340.00'))
        self.assertEqual(medicao.total_faturamento_direto, Decimal('100.00'))
        self.assertEqual(medicao.base_impostos, Decimal('240.00'))
        self.assertEqual(medicao.total_mao_obra_periodo, Decimal('100.00'))

        response = self.client.post(
            reverse('editar_medicao_construtora', args=[medicao.id]),
            {
                'numero': '1',
                'periodo_inicio': '2026-01-01',
                'periodo_fim': '2026-01-31',
                'data_medicao': '2026-01-31',
                'retencao_tecnica': '0',
                'retencao_tecnica_percentual': '0',
                'issqn': '0',
                'issqn_percentual': '5',
                'inss': '0',
                'inss_percentual': '10',
                'desconto_adicional': '0',
                'desconto_adicional_percentual': '0',
                'observacoes': '',
                f'faturamento_direto_{faturamento.id}_percentual': '50',
                'itens-TOTAL_FORMS': '1',
                'itens-INITIAL_FORMS': '1',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-id': str(medicao.itens.get().id),
                'itens-0-quantidade_periodo': '20',
            },
        )

        self.assertRedirects(response, reverse('editar_medicao_construtora', args=[medicao.id]))
        medicao.refresh_from_db()
        faturamento.refresh_from_db()
        self.assertEqual(medicao.total_faturamento_direto, Decimal('50.00'))
        self.assertEqual(medicao.issqn, Decimal('14.50'))
        self.assertEqual(medicao.inss, Decimal('10.00'))
        self.assertEqual(medicao.total_liquido, Decimal('265.50'))
        self.assertEqual(faturamento.medicao_desconto, 'Medicao 1 (50.00%)')

    def test_pdf_medicao_construtora_pagina_faturamentos_diretos_extensos(self):
        orcamento, item = self._orcamento()
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item, quantidade_periodo=Decimal('20'))
        for index in range(12):
            faturamento = FaturamentoDireto.objects.create(
                obra=self.obra,
                numero_nf=f'FD-{index + 1}',
                empresa_comprou='Cliente',
                valor_nota=Decimal('100.00'),
                descricao=f'Material comprado direto {index + 1}',
                vencimento_boleto='30 dias',
            )
            FaturamentoDiretoMedicao.objects.create(medicao=medicao, faturamento_direto=faturamento)

        response = self.client.get(reverse('medicao_construtora_pdf', args=[medicao.id]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertTrue(response.content.startswith(b'%PDF'))

    def test_percentuais_sao_calculados_mesmo_sem_valor_salvo(self):
        orcamento, item = self._orcamento()
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
            issqn_percentual=Decimal('3.00'),
            inss_percentual=Decimal('1.00'),
            retencao_tecnica_percentual=Decimal('5.00'),
            desconto_adicional_percentual=Decimal('2.00'),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item, quantidade_periodo=Decimal('10'))

        self.assertEqual(medicao.subtotal_periodo, Decimal('170.00'))
        self.assertEqual(medicao.base_impostos, Decimal('170.00'))
        self.assertEqual(medicao.total_mao_obra_periodo, Decimal('50.00'))
        self.assertEqual(medicao.issqn_calculado, Decimal('5.10'))
        self.assertEqual(medicao.inss_calculado, Decimal('0.50'))
        self.assertEqual(medicao.retencao_tecnica_calculada, Decimal('8.50'))
        self.assertEqual(medicao.desconto_adicional_calculado, Decimal('3.40'))
        self.assertEqual(medicao.total_liquido, Decimal('152.50'))

    def test_desconto_adicional_pode_reduzir_base_da_nf(self):
        orcamento, item = self._orcamento()
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
            issqn_percentual=Decimal('5.00'),
            inss_percentual=Decimal('11.00'),
            desconto_adicional=Decimal('40.00'),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item, quantidade_periodo=Decimal('10'))

        self.assertEqual(medicao.subtotal_periodo, Decimal('170.00'))
        self.assertEqual(medicao.base_impostos, Decimal('170.00'))
        self.assertEqual(medicao.base_inss, Decimal('50.00'))
        self.assertEqual(medicao.total_material_periodo, Decimal('100.00'))
        self.assertEqual(medicao.total_mao_obra_periodo, Decimal('50.00'))
        self.assertEqual(medicao.total_equipamentos_periodo, Decimal('20.00'))
        self.assertEqual(medicao.issqn_calculado, Decimal('8.50'))
        self.assertEqual(medicao.inss_calculado, Decimal('5.50'))

        medicao.desconto_adicional_reduz_base_nf = True
        medicao.save(update_fields=['desconto_adicional_reduz_base_nf', 'updated_at'])

        self.assertEqual(medicao.base_impostos, Decimal('130.00'))
        self.assertEqual(medicao.base_inss, Decimal('38.24'))
        self.assertEqual(medicao.valor_material_nf, Decimal('76.47'))
        self.assertEqual(medicao.valor_mao_obra_nf, Decimal('38.24'))
        self.assertEqual(medicao.valor_equipamentos_nf, Decimal('15.29'))
        self.assertEqual(medicao.issqn_calculado, Decimal('6.50'))
        self.assertEqual(medicao.inss_calculado, Decimal('4.21'))

    def test_exclui_medicao_construtora(self):
        orcamento, item = self._orcamento()
        medicao = MedicaoConstrutora.objects.create(
            orcamento=orcamento,
            numero=1,
            periodo_inicio=date(2026, 1, 1),
            periodo_fim=date(2026, 1, 31),
            data_medicao=date(2026, 1, 31),
        )
        ItemMedicaoConstrutora.objects.create(medicao=medicao, item_orcamento=item, quantidade_periodo=Decimal('20'))

        response = self.client.post(reverse('excluir_medicao_construtora', args=[medicao.id]))

        self.assertRedirects(response, reverse('detalhe_orcamento_medicao', args=[orcamento.id]))
        self.assertFalse(MedicaoConstrutora.objects.filter(id=medicao.id).exists())
        self.assertTrue(OrcamentoMedicao.objects.filter(id=orcamento.id).exists())

    def test_medicao_empreiteiro_simples_e_exportacoes(self):
        medicao = MedicaoEmpreiteiro.objects.create(
            empresa=self.empresa,
            tipo=MedicaoEmpreiteiro.TIPO_SIMPLES,
            obra=self.obra,
            empreiteiro='Empreiteiro',
            cpf_cnpj='00.000.000/0001-00',
            pix='pix@teste.com',
            numero=1,
            periodo_inicio=date(2026, 3, 1),
            periodo_fim=date(2026, 3, 31),
            data_medicao=date(2026, 3, 31),
            retencao_tecnica=Decimal('15.00'),
        )
        ItemMedicaoEmpreiteiro.objects.create(
            medicao=medicao,
            item='1',
            descricao='Servico simples',
            unidade='un',
            quantidade_periodo=Decimal('2'),
            valor_unitario=Decimal('100.00'),
        )

        self.assertEqual(medicao.total_liquido, Decimal('185.00'))
        pdf = self.client.get(reverse('medicao_empreiteiro_pdf', args=[medicao.id]))
        excel = self.client.get(reverse('medicao_empreiteiro_excel', args=[medicao.id]))
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf['Content-Type'], 'application/pdf')
        self.assertEqual(excel.status_code, 200)
        self.assertIn('spreadsheetml', excel['Content-Type'])

    def test_medicao_simples_usa_cadastro_de_empreiteiro(self):
        empreiteiro = Empreiteiro.objects.create(
            empresa=self.empresa,
            nome='Empreiteiro cadastrado',
            cpf_cnpj='11.111.111/0001-11',
            pix='pix@empreiteiro.com',
        )

        response = self.client.post(
            reverse('nova_medicao_empreiteiro_simples'),
            {
                'obra': self.obra.id,
                'empreiteiro_cadastro': empreiteiro.id,
                'empreiteiro': '',
                'cpf_cnpj': '',
                'pix': '',
                'numero': '1',
                'periodo_inicio': '2026-03-01',
                'periodo_fim': '2026-03-31',
                'data_medicao': '2026-03-31',
                'observacoes': '',
                'itens-TOTAL_FORMS': '1',
                'itens-INITIAL_FORMS': '0',
                'itens-MIN_NUM_FORMS': '0',
                'itens-MAX_NUM_FORMS': '1000',
                'itens-0-item_orcamento': '',
                'itens-0-item': '1',
                'itens-0-descricao': 'Servico',
                'itens-0-unidade': 'un',
                'itens-0-quantidade_periodo': '1',
                'itens-0-valor_unitario': '100.00',
            },
        )

        medicao = MedicaoEmpreiteiro.objects.get()
        self.assertRedirects(response, reverse('editar_medicao_empreiteiro', args=[medicao.id]))
        self.assertEqual(medicao.empreiteiro_cadastro, empreiteiro)
        self.assertEqual(medicao.empreiteiro, 'Empreiteiro cadastrado')
        self.assertEqual(medicao.cpf_cnpj, '11.111.111/0001-11')
        self.assertEqual(medicao.pix, 'pix@empreiteiro.com')

    def test_medicao_sem_empreiteiro_cadastrado_nao_salva(self):
        orcamento, item = self._orcamento()
        orcamento.tipo = OrcamentoMedicao.TIPO_EMPREITEIRO
        orcamento.save(update_fields=['tipo'])

        response = self.client.post(
            reverse('nova_medicao_empreiteiro_cumulativa', args=[orcamento.id]),
            {
                'obra': self.obra.id,
                'empreiteiro_cadastro': '',
                'empreiteiro': 'Novo Empreiteiro',
                'cpf_cnpj': '22.222.222/0001-22',
                'pix': 'pix@novo.com',
                'numero': '1',
                'periodo_inicio': '2026-04-01',
                'periodo_fim': '2026-04-30',
                'data_medicao': '2026-04-30',
                'observacoes': '',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertFalse(MedicaoEmpreiteiro.objects.filter(orcamento=orcamento).exists())
        self.assertFalse(Empreiteiro.objects.filter(nome='Novo Empreiteiro').exists())
        self.assertContains(response, 'Selecione um empreiteiro cadastrado')

    def test_medicao_cumulativa_reaproveita_empreiteiro_cadastrado(self):
        empreiteiro = Empreiteiro.objects.create(
            empresa=self.empresa,
            nome='Novo Empreiteiro',
            cpf_cnpj='22.222.222/0001-22',
            pix='pix@novo.com',
        )
        orcamento, item = self._orcamento()
        orcamento.tipo = OrcamentoMedicao.TIPO_EMPREITEIRO
        orcamento.save(update_fields=['tipo'])

        primeira = MedicaoEmpreiteiro.objects.create(
            empresa=self.empresa,
            tipo=MedicaoEmpreiteiro.TIPO_CUMULATIVA,
            orcamento=orcamento,
            obra=self.obra,
            empreiteiro_cadastro=empreiteiro,
            empreiteiro=empreiteiro.nome,
            cpf_cnpj=empreiteiro.cpf_cnpj,
            pix=empreiteiro.pix,
            numero=1,
            periodo_inicio=date(2026, 4, 1),
            periodo_fim=date(2026, 4, 30),
            data_medicao=date(2026, 4, 30),
        )
        ItemMedicaoEmpreiteiro.objects.create(
            medicao=primeira,
            item_orcamento=item,
            quantidade_periodo=Decimal('1'),
        )

        response = self.client.get(reverse('nova_medicao_empreiteiro_cumulativa', args=[orcamento.id]))

        self.assertContains(response, 'Novo Empreiteiro')
        self.assertContains(response, 'pix@novo.com')

    def test_exclui_medicao_empreiteiro_simples(self):
        medicao = MedicaoEmpreiteiro.objects.create(
            empresa=self.empresa,
            tipo=MedicaoEmpreiteiro.TIPO_SIMPLES,
            obra=self.obra,
            empreiteiro='Empreiteiro',
            numero=1,
            periodo_inicio=date(2026, 3, 1),
            periodo_fim=date(2026, 3, 31),
            data_medicao=date(2026, 3, 31),
        )
        ItemMedicaoEmpreiteiro.objects.create(
            medicao=medicao,
            item='1',
            descricao='Servico simples',
            unidade='un',
            quantidade_periodo=Decimal('2'),
            valor_unitario=Decimal('100.00'),
        )

        response = self.client.post(reverse('excluir_medicao_empreiteiro', args=[medicao.id]))

        self.assertRedirects(response, reverse('medicoes_empreiteiros_home'))
        self.assertFalse(MedicaoEmpreiteiro.objects.filter(id=medicao.id).exists())
