import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from controles.models import EquipamentoLocadoCatalogo, LocacaoEquipamento, LocadoraEquipamento
from obras.models import Obra

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - protected by requirements
    raise CommandError('openpyxl nao esta instalado. Adicione a dependencia antes de importar locacoes.') from exc


STATUS_MAP = {
    'na obra': 'locado',
    'obra': 'locado',
    'ag coleta': 'retirada_solicitada',
    'coletado': 'retirado',
    'recolhida': 'retirado',
    'ag entrega': 'aguardando_entrega',
}

IGNORED_SHEETS = {'roubados'}

HEADER_ALIASES = {
    'data_locacao': {'data aluguel', 'data', 'alugado'},
    'solicitante': {'quem locou', 'locou'},
    'locadora': {'locadora'},
    'status': {'situacao'},
    'equipamento': {'equipamento'},
    'contrato': {'contrato'},
    'quantidade': {'qntd'},
    'data_retirada': {'data coleta', 'data devolucao', 'coleta', 'devolucao'},
    'observacoes': {'observacao'},
    'prazo': {'prazo'},
    'valor': {'valor', 'valor mensal'},
    'obra': {'obra'},
}


def normalize_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    text = text.lower()
    return re.sub(r'\s+', ' ', text)


def canonical_obra_name(sheet_name):
    return re.sub(r'\s*\(.*?\)\s*', '', sheet_name).strip()


def parse_status(raw_status):
    return STATUS_MAP.get(normalize_text(raw_status), 'locado')


def parse_date(value):
    if isinstance(value, datetime):
        if 2000 <= value.year <= 2100:
            return value.date()
        return None
    if isinstance(value, date):
        if 2000 <= value.year <= 2100:
            return value
        return None
    if isinstance(value, str):
        text = value.strip()
        for date_format in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(text, date_format).date()
            except ValueError:
                continue
    return None


def stringify_contract(value):
    if value in (None, '', '-'):
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m')
    if isinstance(value, date):
        return value.strftime('%Y-%m')
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_int(value):
    if value in (None, ''):
        return 1
    try:
        return max(int(Decimal(str(value).replace(',', '.'))), 1)
    except (InvalidOperation, ValueError):
        return 1


def parse_decimal(value):
    if value in (None, ''):
        return Decimal('0')
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value)).quantize(Decimal('0.01'))

    numbers = re.findall(r'\d+[.,]?\d*', str(value))
    if not numbers:
        return Decimal('0')

    cleaned = [Decimal(item.replace('.', '').replace(',', '.')) for item in numbers]
    return max(cleaned).quantize(Decimal('0.01'))


class Command(BaseCommand):
    help = 'Importa locacoes de equipamentos a partir de uma planilha Excel.'

    def add_arguments(self, parser):
        parser.add_argument('arquivo', help='Caminho do arquivo .xlsx')

    def map_headers(self, row):
        mapping = {}
        for index, value in enumerate(row):
            normalized = normalize_text(value)
            for field, aliases in HEADER_ALIASES.items():
                if normalized in aliases:
                    mapping[field] = index
        return mapping

    def find_header_row(self, worksheet):
        for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            mapping = self.map_headers(row)
            if 'equipamento' in mapping and 'locadora' in mapping:
                return row_index, mapping
        return None, None

    def get_obra(self, sheet_name):
        if normalize_text(sheet_name) in IGNORED_SHEETS:
            return None

        target_name = canonical_obra_name(sheet_name)
        target_key = normalize_text(target_name)

        for obra in Obra.objects.all():
            obra_key = normalize_text(obra.nome_obra)
            if target_key in obra_key or obra_key in target_key:
                return obra

        return Obra.objects.create(nome_obra=target_name)

    def row_value(self, row, mapping, key, previous=None):
        index = mapping.get(key)
        if index is None or index >= len(row):
            return previous
        value = row[index]
        if value in (None, ''):
            return previous
        return value

    def import_sheet(self, worksheet):
        obra = self.get_obra(worksheet.title)
        if obra is None:
            return 0

        header_row, mapping = self.find_header_row(worksheet)
        if not header_row:
            return 0

        imported = 0
        previous_context = {}

        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            raw_equipment = self.row_value(row, mapping, 'equipamento')
            if not raw_equipment:
                continue

            equipment_name = str(raw_equipment).strip()
            if equipment_name.upper() == 'EQUIPAMENTO':
                continue

            context = {
                'data_locacao': self.row_value(row, mapping, 'data_locacao', previous_context.get('data_locacao')),
                'solicitante': self.row_value(row, mapping, 'solicitante', previous_context.get('solicitante')),
                'locadora': self.row_value(row, mapping, 'locadora', previous_context.get('locadora')),
                'status': self.row_value(row, mapping, 'status', previous_context.get('status')),
                'contrato': self.row_value(row, mapping, 'contrato', previous_context.get('contrato')),
                'quantidade': self.row_value(row, mapping, 'quantidade', previous_context.get('quantidade')),
                'data_retirada': self.row_value(row, mapping, 'data_retirada', previous_context.get('data_retirada')),
                'observacoes': self.row_value(row, mapping, 'observacoes', previous_context.get('observacoes')),
                'prazo': self.row_value(row, mapping, 'prazo', previous_context.get('prazo')),
                'valor': self.row_value(row, mapping, 'valor', previous_context.get('valor')),
            }

            if not context['locadora']:
                continue

            previous_context = context

            locadora, _ = LocadoraEquipamento.objects.get_or_create(nome=str(context['locadora']).strip())
            equipamento, _ = EquipamentoLocadoCatalogo.objects.get_or_create(nome=equipment_name)

            observacoes = []
            if context.get('observacoes'):
                observacoes.append(str(context['observacoes']).strip())

            data_retirada = parse_date(context.get('data_retirada'))
            if context.get('data_retirada') and data_retirada is None:
                observacoes.append(f'Data coleta original: {context["data_retirada"]}')

            prazo = ''
            if context.get('prazo'):
                prazo_date = parse_date(context['prazo'])
                prazo = prazo_date.strftime('%d/%m/%Y') if prazo_date else str(context['prazo']).strip()

            if context.get('valor') not in (None, '') and isinstance(context['valor'], str):
                value_text = str(context['valor']).strip()
                if len(re.findall(r'\d+[.,]?\d*', value_text)) > 1:
                    observacoes.append(f'Valor original informado: {value_text}')

            defaults = {
                'status': parse_status(context.get('status')),
                'data_retirada': data_retirada,
                'prazo': prazo,
                'valor_referencia': parse_decimal(context.get('valor')),
                'observacoes': ' | '.join(item for item in observacoes if item),
            }

            data_locacao = parse_date(context.get('data_locacao')) or date.today()
            solicitante = str(context.get('solicitante') or '').strip()
            numero_contrato = stringify_contract(context.get('contrato'))
            quantidade = parse_int(context.get('quantidade'))

            LocacaoEquipamento.objects.update_or_create(
                obra=obra,
                equipamento=equipamento,
                locadora=locadora,
                data_locacao=data_locacao,
                solicitante=solicitante,
                numero_contrato=numero_contrato,
                quantidade=quantidade,
                defaults=defaults,
            )
            imported += 1

        return imported

    @transaction.atomic
    def handle(self, *args, **options):
        workbook = load_workbook(options['arquivo'], data_only=True)
        imported = 0

        for worksheet in workbook.worksheets:
            imported += self.import_sheet(worksheet)

        self.stdout.write(self.style.SUCCESS(f'{imported} locacoes importadas/atualizadas com sucesso.'))
