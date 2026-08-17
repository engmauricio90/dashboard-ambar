from dataclasses import dataclass
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from documentos.formatting import format_date_br
from documentos.theme import DocumentTheme


@dataclass(frozen=True)
class ExcelColumn:
    key: str
    label: str
    width: int = 16
    align: str = 'left'
    number_format: str | None = None
    wrap: bool = False


class ExcelReportBuilder:
    MONEY_FORMAT = '"R$" #,##0.00'
    DATE_FORMAT = 'dd/mm/yyyy'
    DECIMAL_FORMAT = '#,##0.00'
    PERCENT_FORMAT = '0.00%'

    FILLS = {
        'header': 'E8EEF3',
        'zebra': 'FAFBFD',
        'surface': 'F3F4F6',
        'contract': 'E8EEF7',
        'measured': 'DCF5E5',
        'receivable': 'FEE2E2',
    }

    def __init__(self, empresa=None, title='', subtitle='', sheet_name='Relatorio', orientation='landscape'):
        self.theme = DocumentTheme(empresa, orientation=orientation)
        self.empresa = empresa
        self.title = title
        self.subtitle = subtitle
        self.orientation = orientation
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name[:31]
        self.current_row = 1
        self._setup_page()

    def _setup_page(self):
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        self.ws.page_setup.orientation = self.orientation
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = 0
        self.ws.sheet_properties.pageSetUpPr.fitToPage = True
        self.ws.page_margins.left = 0.25
        self.ws.page_margins.right = 0.25
        self.ws.page_margins.top = 0.45
        self.ws.page_margins.bottom = 0.45
        company = self.theme.company_name
        self.ws.oddFooter.center.text = company
        self.ws.oddFooter.right.text = 'Pagina &P de &N'

    def add_header(self, emitted_on=None, filters=None):
        self.ws.cell(self.current_row, 1, self.title)
        self.ws.cell(self.current_row, 1).font = Font(bold=True, size=16, color='111827')
        self.current_row += 1
        self.ws.cell(self.current_row, 1, self.theme.company_name)
        self.ws.cell(self.current_row, 1).font = Font(bold=True, size=11, color='4B5563')
        if emitted_on:
            self.ws.cell(self.current_row, 3, f'Emitido em {format_date_br(emitted_on)}')
        self.current_row += 1
        if self.subtitle:
            self.ws.cell(self.current_row, 1, self.subtitle)
            self.current_row += 1
        if filters:
            self.ws.cell(self.current_row, 1, filters)
            self.current_row += 1
        self.current_row += 1

    def _fill(self, color):
        return PatternFill('solid', fgColor=self.FILLS.get(color, color))

    def add_table(self, columns, rows):
        header_row = self.current_row
        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = self._fill('header')
        zebra_fill = self._fill('zebra')
        for col_index, column in enumerate(columns, start=1):
            cell = self.ws.cell(header_row, col_index, column.label)
            cell.font = Font(bold=True, color='111827')
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            self.ws.column_dimensions[get_column_letter(col_index)].width = column.width
        for row in rows:
            self.current_row += 1
            if row.get('__merge_label__'):
                end_col = len(columns)
                self.ws.merge_cells(
                    start_row=self.current_row,
                    start_column=1,
                    end_row=self.current_row,
                    end_column=end_col,
                )
                cell = self.ws.cell(self.current_row, 1, row.get('__merge_label__'))
                cell.fill = self._fill(row.get('__bg') or 'surface')
                cell.font = Font(bold=True, color='111827')
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                for col_index in range(2, end_col + 1):
                    extra = self.ws.cell(self.current_row, col_index)
                    extra.fill = self._fill(row.get('__bg') or 'surface')
                    extra.border = border
                continue
            is_zebra = (self.current_row - header_row) % 2 == 0
            row_fill = self._fill(row.get('__bg')) if row.get('__bg') else None
            row_font = Font(bold=bool(row.get('__bold')), color='111827')
            cell_bgs = row.get('__cell_bgs') if isinstance(row.get('__cell_bgs'), dict) else {}
            for col_index, column in enumerate(columns, start=1):
                cell = self.ws.cell(self.current_row, col_index, row.get(column.key))
                cell.border = border
                if column.key in cell_bgs:
                    cell.fill = self._fill(cell_bgs[column.key])
                elif row_fill:
                    cell.fill = row_fill
                elif is_zebra:
                    cell.fill = zebra_fill
                if row.get('__bold'):
                    cell.font = row_font
                cell.alignment = Alignment(horizontal=column.align, vertical='center', wrap_text=column.wrap)
                if column.number_format:
                    cell.number_format = column.number_format
        last_col = get_column_letter(len(columns))
        self.ws.freeze_panes = self.ws.cell(header_row + 1, 1).coordinate
        self.ws.auto_filter.ref = f'A{header_row}:{last_col}{self.current_row}'
        self.ws.print_title_rows = f'{header_row}:{header_row}'
        self.current_row += 2

    def build(self):
        output = BytesIO()
        self.wb.save(output)
        return output.getvalue()

    def response(self, filename):
        response = HttpResponse(
            self.build(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
